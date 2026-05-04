"""
Extract unique source texts from a target locale's xlsx files.

Sources scanned (all under Trans To Vostok/<target_locale>/):
    - Translation.xlsx   (game text, with method/untranslatable filters)
    - Texture.xlsx       (image labels — capitalized Text/Translation columns)
    - Glossary.xlsx      (translator reference, untranslatable filter only)

Filter logic per row (per file's column conventions):
    - text must be non-empty
    - translation must be empty (already-translated rows are excluded — saves
      DeepL quota on re-runs and preserves human/curated edits)
    - method != ignore  (operational exclusion, only for Translation.xlsx)
    - method != pattern (regex source, only for Translation.xlsx)
    - untranslatable != 1 (only when the file has that column)

Then dedupe by exact text (same English source across all files -> single
unique entry, ensuring identical DeepL output everywhere).

Output (under <mod_root>/.tmp/unique_text/<target_locale>/):
    unique.tsv     unique_id, text, occurrences, char_count
    mapping.tsv    unique_id, source_file, sheet, row_in_sheet, text
    stats.txt      human-readable summary

Usage:
    python tools/utils/export_unique_text.py <target_locale>

Example:
    python tools/utils/export_unique_text.py French
    python tools/utils/export_unique_text.py Japanese
"""
import csv
import sys
from pathlib import Path

try:
    import openpyxl  # noqa: F401
except ImportError:
    print("ERROR: openpyxl is required. pip install -r tools/requirements.txt", file=sys.stderr)
    sys.exit(1)

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass

try:
    import openpyxl as _openpyxl  # noqa: F401
except ImportError:
    pass  # already checked above


BOOL_TRUE = {"1", "true"}

# methods that should NOT be sent to a translator (Translation.xlsx only)
EXCLUDED_METHODS = {"ignore", "pattern"}

# Per-xlsx column configuration. None == column not present in this file.
XLSX_FILES = [
    {
        "name": "Translation.xlsx",
        "text_col": "text",
        "trans_col": "translation",
        "method_col": "method",
        "untrans_col": "untranslatable",
    },
    {
        "name": "Texture.xlsx",
        "text_col": "Text",
        "trans_col": "Translation",
        "method_col": None,
        "untrans_col": None,
    },
    {
        "name": "Glossary.xlsx",
        "text_col": "text",
        "trans_col": "translation",
        "method_col": None,
        "untrans_col": "untranslatable",
    },
]


def _empty_stats() -> dict:
    return {
        "sheets": {},
        "total_data_rows": 0,
        "excluded_empty_text": 0,
        "excluded_already_translated": 0,
        "excluded_method_ignore": 0,
        "excluded_method_pattern": 0,
        "excluded_untranslatable": 0,
        "candidate_rows": 0,
    }


def collect_from_xlsx(xlsx_path: Path, cfg: dict) -> tuple[list[dict], dict]:
    """Read one xlsx using its column config, return (rows, stats)."""
    import openpyxl
    rows: list[dict] = []
    stats = _empty_stats()
    if not xlsx_path.exists():
        return rows, stats

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name == "MetaData":
                continue
            ws = wb[sheet_name]
            iter_rows = ws.iter_rows(values_only=True)
            header = next(iter_rows, None)
            if header is None:
                continue
            header_map = {
                str(h).strip(): i for i, h in enumerate(header) if h is not None
            }

            text_idx = header_map.get(cfg["text_col"])
            trans_idx = header_map.get(cfg["trans_col"])
            if text_idx is None or trans_idx is None:
                continue  # not a translation sheet

            method_idx = header_map.get(cfg["method_col"]) if cfg["method_col"] else None
            untrans_idx = header_map.get(cfg["untrans_col"]) if cfg["untrans_col"] else None

            per_sheet = {"data_rows": 0, "candidates": 0, "candidate_chars": 0}

            for row_idx, row in enumerate(iter_rows, start=1):
                if row is None:
                    continue
                stats["total_data_rows"] += 1
                per_sheet["data_rows"] += 1

                text = row[text_idx] if text_idx < len(row) else None
                if text is None or str(text).strip() == "":
                    stats["excluded_empty_text"] += 1
                    continue
                text = str(text)

                translation = row[trans_idx] if trans_idx < len(row) else None
                if translation is not None and str(translation).strip() != "":
                    stats["excluded_already_translated"] += 1
                    continue

                if untrans_idx is not None:
                    uval = row[untrans_idx] if untrans_idx < len(row) else None
                    if uval is not None and str(uval).strip().lower() in BOOL_TRUE:
                        stats["excluded_untranslatable"] += 1
                        continue

                if method_idx is not None:
                    mval = row[method_idx] if method_idx < len(row) else None
                    method_str = str(mval).strip().lower() if mval is not None else ""
                    if method_str == "ignore":
                        stats["excluded_method_ignore"] += 1
                        continue
                    if method_str == "pattern":
                        stats["excluded_method_pattern"] += 1
                        continue

                rows.append({
                    "source_file": xlsx_path.name,
                    "sheet": sheet_name,
                    "row_in_sheet": row_idx,
                    "text": text,
                })
                stats["candidate_rows"] += 1
                per_sheet["candidates"] += 1
                per_sheet["candidate_chars"] += len(text)

            stats["sheets"][f"{xlsx_path.stem}/{sheet_name}"] = per_sheet
    finally:
        wb.close()

    return rows, stats


def collect_candidates(locale_dir: Path) -> tuple[list[dict], dict]:
    """Read all configured xlsx files and merge candidate rows + stats."""
    all_rows: list[dict] = []
    combined = _empty_stats()

    for cfg in XLSX_FILES:
        xlsx_path = locale_dir / cfg["name"]
        rows, stats = collect_from_xlsx(xlsx_path, cfg)
        all_rows.extend(rows)
        combined["sheets"].update(stats["sheets"])
        for k in ("total_data_rows", "excluded_empty_text", "excluded_already_translated",
                  "excluded_method_ignore", "excluded_method_pattern",
                  "excluded_untranslatable", "candidate_rows"):
            combined[k] += stats[k]

    return all_rows, combined


def deduplicate(rows: list[dict]) -> tuple[list[dict], list[dict]]:
    """Group rows by exact text. Returns (unique, mapping)."""
    text_to_id: dict[str, int] = {}
    unique: list[dict] = []
    mapping: list[dict] = []

    for row in rows:
        text = row["text"]
        if text in text_to_id:
            uid = text_to_id[text]
            unique[uid - 1]["occurrences"] += 1
        else:
            uid = len(unique) + 1
            text_to_id[text] = uid
            unique.append({
                "unique_id": uid,
                "text": text,
                "occurrences": 1,
                "char_count": len(text),
            })
        mapping.append({
            "unique_id": uid,
            "source_file": row.get("source_file", ""),
            "sheet": row["sheet"],
            "row_in_sheet": row["row_in_sheet"],
            "text": text,
        })

    return unique, mapping


def write_tsv(path: Path, columns: list[str], rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    try:
        with open(tmp, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)
            writer.writerow(columns)
            for row in rows:
                writer.writerow([row.get(c, "") for c in columns])
        tmp.replace(path)
    except Exception:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
        raise


def write_stats(path: Path, stats: dict, unique: list[dict], mapping: list[dict],
                target_locale: str) -> str:
    path.parent.mkdir(parents=True, exist_ok=True)
    candidate_chars = sum(u["char_count"] * u["occurrences"] for u in unique)
    unique_chars = sum(u["char_count"] for u in unique)
    reduction_pct = (1 - unique_chars / candidate_chars) * 100 if candidate_chars else 0.0

    lines = []
    lines.append(f"Target locale: {target_locale}")
    lines.append("")
    lines.append(f"Total data rows scanned       : {stats['total_data_rows']:>8d}")
    lines.append(f"  excluded (empty text)       : {stats['excluded_empty_text']:>8d}")
    lines.append(f"  excluded (already translated): {stats['excluded_already_translated']:>8d}")
    lines.append(f"  excluded (method=ignore)    : {stats['excluded_method_ignore']:>8d}")
    lines.append(f"  excluded (method=pattern)   : {stats['excluded_method_pattern']:>8d}")
    lines.append(f"  excluded (untranslatable=1) : {stats['excluded_untranslatable']:>8d}")
    lines.append(f"  candidate rows              : {stats['candidate_rows']:>8d}")
    lines.append("")
    lines.append(f"Unique texts                  : {len(unique):>8d}")
    lines.append(f"Mapping entries               : {len(mapping):>8d}")
    lines.append("")
    lines.append(f"Total chars (all candidates)  : {candidate_chars:>8d}")
    lines.append(f"Total chars (unique only)     : {unique_chars:>8d}")
    lines.append(f"Dedup char reduction          : {reduction_pct:>7.2f} %")
    lines.append("")
    lines.append("Per-sheet breakdown:")
    lines.append(f"  {'sheet':<22s}  {'data':>6s}  {'cand':>6s}  {'chars':>8s}")
    for sheet_name, per in stats["sheets"].items():
        lines.append(
            f"  {sheet_name:<22s}  {per['data_rows']:>6d}  "
            f"{per['candidates']:>6d}  {per['candidate_chars']:>8d}"
        )

    summary = "\n".join(lines)
    path.write_text(summary + "\n", encoding="utf-8")
    return summary


def main() -> int:
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    if not args:
        print("Usage: python tools/utils/export_unique_text.py <target_locale>")
        print("Example: python tools/utils/export_unique_text.py French")
        return 1

    target_locale = args[0]

    script_dir = Path(__file__).resolve().parent
    # script_dir = mods/Trans To Vostok/tools/utils
    mod_root = script_dir.parent.parent
    pkg_root = mod_root / "Trans To Vostok"
    locale_dir = pkg_root / target_locale
    out_dir = mod_root / ".tmp" / "unique_text" / target_locale

    if not locale_dir.exists():
        print(f"[ERROR] Locale folder not found: {locale_dir}")
        print(f"  Create it first by copying from Template.")
        return 1

    # Report which xlsx files will be scanned
    print(f"Locale folder : {locale_dir}")
    print(f"Output folder : {out_dir}")
    print(f"Scanning xlsx files:")
    for cfg in XLSX_FILES:
        p = locale_dir / cfg["name"]
        marker = "yes" if p.exists() else "absent (skip)"
        print(f"  - {cfg['name']:<20s}  ({marker})")
    print()

    print("[1/3] Scanning sheets...")
    try:
        rows, stats = collect_candidates(locale_dir)
    except PermissionError as e:
        print(f"[ERROR] Cannot read xlsx (file locked? close Excel and retry): {e}")
        return 1

    print(f"  -> {stats['candidate_rows']} candidate rows from {stats['total_data_rows']} total")
    print()

    print("[2/3] Deduplicating...")
    unique, mapping = deduplicate(rows)
    print(f"  -> {len(unique)} unique texts (from {len(mapping)} candidates)")
    print()

    print("[3/3] Writing output...")
    unique_path = out_dir / "unique.tsv"
    mapping_path = out_dir / "mapping.tsv"
    stats_path = out_dir / "stats.txt"

    write_tsv(unique_path, ["unique_id", "text", "occurrences", "char_count"], unique)
    write_tsv(mapping_path, ["unique_id", "source_file", "sheet", "row_in_sheet", "text"], mapping)
    summary = write_stats(stats_path, stats, unique, mapping, target_locale)

    print(f"  -> {unique_path.relative_to(mod_root)}")
    print(f"  -> {mapping_path.relative_to(mod_root)}")
    print(f"  -> {stats_path.relative_to(mod_root)}")
    print()
    print("=" * 60)
    print(summary)
    return 0


if __name__ == "__main__":
    sys.exit(main())
