"""Rebuild <locale>/Translation.xlsx from canonical TSVs.

Source : <project_root>/Translation_TSV/<locale>/Translation/*.tsv
Output : <pkg_root>/<locale>/Translation.xlsx (overwritten)

Sheet ordering : `_sheet_order.txt` next to the TSVs.
Column widths  : unified policy in `tools/width.json`, keyed by HEADER NAME
                 (not column letter). Sheet "MetaData" uses the "MetaData"
                 entry; every other sheet uses the "Translation" entry.
                 Header names absent from the sheet are silently skipped.
Conditional formatting (Translation-category sheets only, MetaData skipped):
  - untranslatable = 1                            -> red bg / red text
  - method = substr/static/ignore/literal/Pattern -> green/blue/red/yellow/yellow,red
  - duplicate values in unique_id/text/translation -> red bg / red text
Group separators (Translation-category sheets only, MetaData skipped):
  - WHERE change between consecutive rows -> thick line (precedence)
  - SUB change between consecutive rows   -> thin line
  - End of data (last data row -> first empty row) -> thick line

Usage:
    python tools/utils/rebuild_translation_xlsx.py <locale>
"""
import sys, io, csv, json
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.utils import get_column_letter

SCRIPT_DIR = Path(__file__).resolve().parent           # tools/utils
PROJECT_ROOT = SCRIPT_DIR.parent.parent                 # mods/Trans To Vostok
TSV_ROOT = PROJECT_ROOT / "Translation_TSV"
PKG_ROOT = PROJECT_ROOT / "Trans To Vostok"
WIDTH_POLICY = PROJECT_ROOT / "tools" / "width.json"

CATEGORY = "Translation"
THICK_RIGHT_HEADERS = {"KIND", "untranslatable", "unique_id", "translation"}
WRAP_TEXT_HEADERS = {"text", "translation", "description"}  # case-insensitive
GROUP_THICK_HEADER = "WHERE"
GROUP_THIN_HEADER = "SUB"
DUP_CHECK_HEADERS = ["unique_id", "text", "translation"]
FORMATTED_ROW_LIMIT = 5000

DASH = Side(style="dashed")
THICK = Side(style="thick")
HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
HEADER_FONT = Font(bold=True)

RED_BG, RED_FG       = "FFC7CE", "9C0006"
YELLOW_BG, YELLOW_FG = "FFEB9C", "9C5700"
GREEN_BG, GREEN_FG   = "C6EFCE", "006100"
BLUE_BG, BLUE_FG     = "BDD7EE", "1F4E78"


def sheet_name_for(tsv: Path) -> str:
    name = tsv.stem[:31]
    for bad in '[]:*?/\\':
        name = name.replace(bad, "_")
    return name


def _resolve_tsv_order(src_dir: Path) -> list[Path]:
    order_file = src_dir / "_sheet_order.txt"
    if order_file.exists():
        names = [ln.strip() for ln in order_file.read_text(encoding="utf-8").splitlines()
                 if ln.strip()]
        ordered = [src_dir / f"{name}.tsv" for name in names]
        missing = [p for p in ordered if not p.exists()]
        for p in missing:
            print(f"  [WARN] order references missing TSV: {p.name}")
        ordered = [p for p in ordered if p.exists()]
        listed = {p.name for p in ordered}
        extras = sorted(p for p in src_dir.glob("*.tsv") if p.name not in listed)
        for p in extras:
            print(f"  [INFO] TSV not in order file (appending last): {p.name}")
        return ordered + extras
    return sorted(src_dir.glob("*.tsv"))


def _load_width_policy(policy_path: Path) -> dict[str, dict[str, float]]:
    if not policy_path.exists():
        print(f"  [WARN] width policy missing: {policy_path}")
        return {}
    try:
        data = json.loads(policy_path.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"  [WARN] could not parse {policy_path.name}: {e}")
        return {}
    return {k: v for k, v in data.items() if not k.startswith("_")}


def _widths_for_sheet(sheet_name: str, policy: dict) -> dict[str, float]:
    key = "MetaData" if sheet_name == "MetaData" else CATEGORY
    return policy.get(key, {})


def _cell_is_rule(formula_value: str, bg: str, fg: str) -> CellIsRule:
    return CellIsRule(
        operator="equal",
        formula=[formula_value],
        fill=PatternFill("solid", bgColor=bg),
        font=Font(color=fg),
    )


def _apply_alignment(ws, header: list, max_row: int, visible_cols: list) -> None:
    for c in visible_cols:
        cell = ws.cell(1, c)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    wrap_cols = []
    for c in visible_cols:
        head = str(header[c - 1] if c - 1 < len(header) else "").strip().lower()
        if head in WRAP_TEXT_HEADERS:
            wrap_cols.append(c)
    if not wrap_cols:
        return

    for r in range(2, max_row + 1):
        for c in wrap_cols:
            cell = ws.cell(r, c)
            existing = cell.alignment
            cell.alignment = Alignment(
                horizontal=existing.horizontal,
                vertical=existing.vertical,
                wrap_text=True,
            )


def _apply_group_separators(ws, header: list, n_rows: int, visible_cols: list) -> None:
    """GROUP_THICK_HEADER change -> thick (precedence). GROUP_THIN_HEADER
    change -> thin. Always emits a thick end-of-data marker after the last row.
    """
    thick_idx = None
    thin_idx = None
    for i, h in enumerate(header):
        if not h:
            continue
        name = str(h).strip()
        if name == GROUP_THICK_HEADER:
            thick_idx = i + 1
        elif name == GROUP_THIN_HEADER:
            thin_idx = i + 1

    if n_rows >= 2:
        end_side = Side(style="thick")
        for c in visible_cols:
            last_cell = ws.cell(n_rows, c)
            le = last_cell.border
            last_cell.border = Border(
                left=le.left, right=le.right,
                top=le.top, bottom=end_side,
            )
            next_cell = ws.cell(n_rows + 1, c)
            ne = next_cell.border
            next_cell.border = Border(
                left=ne.left, right=ne.right,
                top=end_side, bottom=ne.bottom,
            )

    if thick_idx is None and thin_idx is None:
        return

    prev_thick = None
    prev_thin = None
    for r in range(2, n_rows + 1):
        cur_thick = ws.cell(r, thick_idx).value if thick_idx else None
        cur_thin = ws.cell(r, thin_idx).value if thin_idx else None
        if r >= 3:
            style = None
            if thick_idx and prev_thick != cur_thick:
                style = "thick"
            elif thin_idx and prev_thin != cur_thin:
                style = "thin"
            if style is not None:
                new_side = Side(style=style)
                for c in visible_cols:
                    prev_cell = ws.cell(r - 1, c)
                    pe = prev_cell.border
                    prev_cell.border = Border(
                        left=pe.left, right=pe.right,
                        top=pe.top, bottom=new_side,
                    )
                    cur_cell = ws.cell(r, c)
                    ce = cur_cell.border
                    cur_cell.border = Border(
                        left=ce.left, right=ce.right,
                        top=new_side, bottom=ce.bottom,
                    )
        prev_thick = cur_thick
        prev_thin = cur_thin


def _apply_conditional_formatting(ws, header: list, max_row: int) -> None:
    name_to_letter = {}
    for i, name in enumerate(header, 1):
        if name and str(name).strip():
            name_to_letter[str(name).strip()] = get_column_letter(i)

    def col_letter(name: str) -> str | None:
        return name_to_letter.get(name)

    end_row = max(2, max_row)

    col = col_letter("untranslatable")
    if col:
        ws.conditional_formatting.add(
            f"{col}2:{col}{end_row}",
            _cell_is_rule('"1"', RED_BG, RED_FG),
        )
    method_col = col_letter("method")
    if method_col:
        rng = f"{method_col}2:{method_col}{end_row}"
        for value, bg, fg in [
            ("substr",  GREEN_BG,  GREEN_FG),
            ("static",  BLUE_BG,   BLUE_FG),
            ("ignore",  RED_BG,    RED_FG),
            ("literal", YELLOW_BG, YELLOW_FG),
            ("Pattern", YELLOW_BG, RED_FG),
        ]:
            ws.conditional_formatting.add(rng, _cell_is_rule(f'"{value}"', bg, fg))

    dup_fill = PatternFill("solid", bgColor=RED_BG)
    dup_font = Font(color=RED_FG)
    for header_name in DUP_CHECK_HEADERS:
        col = col_letter(header_name)
        if not col:
            continue
        rng = f"{col}2:{col}{end_row}"
        formula = f'AND({col}2<>"",COUNTIF(${col}$2:${col}${end_row},{col}2)>1)'
        ws.conditional_formatting.add(
            rng,
            FormulaRule(formula=[formula], fill=dup_fill, font=dup_font),
        )


def build(locale: str) -> int:
    src_dir = TSV_ROOT / locale / CATEGORY
    dst = PKG_ROOT / locale / f"{CATEGORY}.xlsx"

    if not src_dir.exists():
        print(f"[SKIP] {locale}/{CATEGORY}: TSV dir not found ({src_dir})")
        return 0

    lock = dst.parent / f"~${dst.name}"
    if lock.exists():
        print(f"  [WARN] Excel lock file detected: {lock.name}")
        print(f"         {dst.name} appears to be open in Excel — write may fail.")

    wb = Workbook()
    wb.remove(wb.active)
    tsv_files = _resolve_tsv_order(src_dir)
    print(f"Source : {src_dir}")
    print(f"Output : {dst}")
    print(f"Found {len(tsv_files)} TSV files")

    width_policy = _load_width_policy(WIDTH_POLICY)

    for tsv in tsv_files:
        sheet_name = sheet_name_for(tsv)
        ws = wb.create_sheet(sheet_name)

        with open(tsv, encoding="utf-8") as f:
            rows = list(csv.reader(f, delimiter="\t"))
        if not rows:
            continue

        for r, row in enumerate(rows, 1):
            for c, val in enumerate(row, 1):
                ws.cell(r, c).value = val

        n_rows = len(rows)
        n_cols = max(len(r) for r in rows)
        header = rows[0]

        non_empty_cols = []
        for c in range(1, n_cols + 1):
            head_val = header[c - 1] if c - 1 < len(header) else ""
            if head_val and str(head_val).strip():
                non_empty_cols.append(c)
        if not non_empty_cols:
            continue

        thick_right_set = set()
        for c in non_empty_cols:
            head = str(header[c - 1] if c - 1 < len(header) else "").strip()
            if head in THICK_RIGHT_HEADERS:
                thick_right_set.add(c)
        first_col = non_empty_cols[0]
        last_col = non_empty_cols[-1]
        thick_right_set.add(last_col)

        thick_left_set = {first_col}
        for c in thick_right_set:
            try:
                idx = non_empty_cols.index(c)
                if idx + 1 < len(non_empty_cols):
                    thick_left_set.add(non_empty_cols[idx + 1])
            except ValueError:
                pass

        max_row_border = max(n_rows, FORMATTED_ROW_LIMIT)

        for r in range(1, max_row_border + 1):
            for c in non_empty_cols:
                left = THICK if c in thick_left_set else DASH
                right = THICK if c in thick_right_set else DASH
                top = DASH
                bottom = DASH

                if r == 1:
                    top = THICK
                    bottom = THICK
                elif r == 2:
                    top = THICK

                cell = ws.cell(r, c)
                cell.border = Border(left=left, right=right, top=top, bottom=bottom)

                if r == 1:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT

        widths = _widths_for_sheet(sheet_name, width_policy)
        if widths:
            name_to_letter = {}
            for i, h in enumerate(header):
                if h and str(h).strip():
                    name_to_letter[str(h).strip()] = get_column_letter(i + 1)
            for header_name, width in widths.items():
                letter = name_to_letter.get(header_name)
                if letter:
                    ws.column_dimensions[letter].width = float(width)

        _apply_conditional_formatting(ws, header, max_row_border)
        if sheet_name != "MetaData":
            _apply_group_separators(ws, header, n_rows, non_empty_cols)
        _apply_alignment(ws, header, max_row_border, non_empty_cols)

        print(f"  {sheet_name}: {n_rows} rows x {len(non_empty_cols)} cols (visible)")

    dst.parent.mkdir(parents=True, exist_ok=True)
    try:
        wb.save(dst)
    except PermissionError:
        print(f"\n[ERROR] Cannot write {dst}")
        print(f"        Close {dst.name} in Excel and re-run.")
        return 1
    print(f"\nWrote: {dst}")
    print(f"Sheets: {wb.sheetnames}")
    return 0


def main(argv: list[str]) -> int:
    if len(argv) < 2:
        print(f"Usage: python {Path(__file__).name} <locale>")
        return 1
    return build(argv[1])


if __name__ == "__main__":
    sys.exit(main(sys.argv))
