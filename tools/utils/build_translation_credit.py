"""
Auto-generate Translation_Credit.md for a locale from xlsx data sources.

Sources:
    Translation.xlsx MetaData sheet:
        - Field "Translator"             -> lead translator(s)
        - Field "Contributor (Translate)" -> translation contributors
    Texture.xlsx (all sheets except MetaData):
        - Column "Reworked by"  -> texture reworkers (unique)
        - Column "Contributors" -> texture contributors (unique)

Multi-line cells (Alt+Enter in Excel) are split into individual names.
Names are deduplicated; lead translators are not repeated in contributor lists.

Output:
    <pkg_root>/<locale>/Translation_Credit.md

Usage:
    python tools/utils/build_translation_credit.py                  # default (Korean)
    python tools/utils/build_translation_credit.py --locale Korean
"""
import argparse
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install -r tools/requirements.txt", file=sys.stderr)
    sys.exit(1)

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass


METADATA_TRANSLATOR_FIELD = "Translator"
METADATA_CONTRIBUTOR_FIELD = "Contributor (Translate)"

TEXTURE_REWORKED_COLUMN = "Reworked by"
TEXTURE_CONTRIBUTORS_COLUMN = "Contributors"


def split_names(value) -> list[str]:
    """Split a possibly multi-line cell value into a list of trimmed names."""
    if value is None:
        return []
    names = []
    for line in str(value).split("\n"):
        line = line.strip()
        if line:
            names.append(line)
    return names


def read_metadata_field(xlsx_path: Path, field_name: str) -> list[str]:
    """Return the value list for a specific field in the MetaData sheet."""
    if not xlsx_path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except (PermissionError, Exception) as e:
        print(f"  [WARN] Cannot read {xlsx_path.name}: {e}")
        return []
    try:
        if "MetaData" not in wb.sheetnames:
            return []
        ws = wb["MetaData"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        try:
            field_idx = header.index("Field")
            value_idx = header.index("Value")
        except ValueError:
            return []
        for row in rows[1:]:
            if row is None or len(row) <= max(field_idx, value_idx):
                continue
            field = str(row[field_idx]).strip() if row[field_idx] is not None else ""
            if field == field_name:
                return split_names(row[value_idx])
        return []
    finally:
        wb.close()


def collect_texture_credits(xlsx_path: Path) -> tuple[set[str], set[str]]:
    """Aggregate unique values for Reworked by / Contributors across all data sheets."""
    reworked: set[str] = set()
    contributors: set[str] = set()
    if not xlsx_path.exists():
        return reworked, contributors
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except (PermissionError, Exception) as e:
        print(f"  [WARN] Cannot read {xlsx_path.name}: {e}")
        return reworked, contributors
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name == "MetaData":
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c).strip() if c is not None else "" for c in rows[0]]
            rb_idx = header.index(TEXTURE_REWORKED_COLUMN) if TEXTURE_REWORKED_COLUMN in header else None
            co_idx = header.index(TEXTURE_CONTRIBUTORS_COLUMN) if TEXTURE_CONTRIBUTORS_COLUMN in header else None
            for row in rows[1:]:
                if row is None:
                    continue
                if rb_idx is not None and rb_idx < len(row):
                    for n in split_names(row[rb_idx]):
                        reworked.add(n)
                if co_idx is not None and co_idx < len(row):
                    for n in split_names(row[co_idx]):
                        contributors.add(n)
    finally:
        wb.close()
    return reworked, contributors


def build_md(
    locale: str,
    lead: list[str],
    translation_contributors: list[str],
    texture_reworked: list[str],
    texture_contributors: list[str],
) -> str:
    def section(title: str, names: list[str]) -> list[str]:
        out = [f"## {title}", ""]
        if names:
            for n in names:
                out.append(f"- {n}")
        else:
            out.append("_(none yet)_")
        out.append("")
        return out

    lines: list[str] = []
    lines.append(f"# {locale} Translation Credits")
    lines.append("")
    lines.append(
        f"People who contributed to translating Road to Vostok into {locale}. "
        f"This includes both text translation and texture / image rework."
    )
    lines.append("")
    lines.extend(section("Lead Translator(s)", lead))
    lines.extend(section("Translation Contributors", translation_contributors))
    lines.extend(section("Texture Reworkers", texture_reworked))
    lines.extend(section("Texture Contributors", texture_contributors))
    lines.append("---")
    lines.append("")
    lines.append(
        "_Auto-generated from `Translation.xlsx` MetaData "
        "(`Translator`, `Contributor (Translate)` fields) "
        "and `Texture.xlsx` (`Reworked by`, `Contributors` columns) "
        "by `tools/utils/build_translation_credit.py`. "
        "Do not edit manually — update the source xlsx files instead._"
    )
    return "\n".join(lines) + "\n"


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate Translation_Credit.md from Translation.xlsx + Texture.xlsx"
    )
    parser.add_argument("--locale", default="Korean", help="Locale folder name (default: Korean)")
    parser.add_argument("--output", default=None, help="Output path (default: <pkg_root>/<locale>/Translation_Credit.md)")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    mod_root = script_dir.parent.parent
    translations_root = mod_root / "Translations"
    pkg_root = mod_root / "Trans To Vostok"
    xlsx_locale_dir = translations_root / args.locale
    output_locale_dir = pkg_root / args.locale
    translation_xlsx = xlsx_locale_dir / "Translation.xlsx"
    texture_xlsx = xlsx_locale_dir / "Texture.xlsx"
    locale_dir = xlsx_locale_dir  # input check below
    output_dir = output_locale_dir

    if not locale_dir.exists():
        print(f"[ERROR] Locale folder not found: {locale_dir}")
        return 1

    if not translation_xlsx.exists() and not texture_xlsx.exists():
        print(f"[SKIP] Neither Translation.xlsx nor Texture.xlsx found in {locale_dir}")
        return 0

    print(f"Locale: {args.locale}")
    print(f"  Translation.xlsx : {'found' if translation_xlsx.exists() else 'missing (skip)'}")
    print(f"  Texture.xlsx     : {'found' if texture_xlsx.exists() else 'missing (skip)'}")
    print()

    lead_list = read_metadata_field(translation_xlsx, METADATA_TRANSLATOR_FIELD)
    translation_contrib_list = read_metadata_field(translation_xlsx, METADATA_CONTRIBUTOR_FIELD)
    texture_reworked_set, texture_contrib_set = collect_texture_credits(texture_xlsx)

    # Dedup policy:
    # - Within MetaData (Translator vs Contributor): mutually exclusive — same person
    #   should be listed only once (as lead, not also as translation contributor).
    # - Texture sections: independent of MetaData — a lead translator who also
    #   did texture work is credited in BOTH lead AND texture section (different roles).
    # - Texture Reworked by vs Contributors: within texture, Reworked by takes
    #   precedence (don't repeat the same name in both texture sections).
    lead_unique: list[str] = []
    seen_meta = set()
    for n in lead_list:
        if n not in seen_meta:
            lead_unique.append(n)
            seen_meta.add(n)

    translation_contrib_unique: list[str] = []
    for n in translation_contrib_list:
        if n not in seen_meta:
            translation_contrib_unique.append(n)
            seen_meta.add(n)

    texture_reworked = sorted(texture_reworked_set)
    texture_contrib = sorted(texture_contrib_set - texture_reworked_set)

    print(f"  Lead translator(s)       : {len(lead_unique)}")
    print(f"  Translation contributors : {len(translation_contrib_unique)}")
    print(f"  Texture reworkers        : {len(texture_reworked)}")
    print(f"  Texture contributors     : {len(texture_contrib)}")
    print()

    output_path = Path(args.output).resolve() if args.output else (output_dir / "Translation_Credit.md")
    md = build_md(
        args.locale, lead_unique, translation_contrib_unique,
        texture_reworked, texture_contrib,
    )
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(md, encoding="utf-8")

    rel = output_path
    try:
        rel = output_path.relative_to(mod_root)
    except ValueError:
        pass
    print(f"Output: {rel}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
