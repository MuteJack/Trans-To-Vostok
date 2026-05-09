"""
Convert xlsx files in language locale folders to canonical TSV files.

This is the xlsx -> TSV side of the editing/VCS workflow:
    xlsx (local working copy, gitignored)  ->  TSV (canonical, committed)

Path mapping:
    <project_root>/Translations/<language>/<file>.xlsx
        ->
    <project_root>/Translations/<language>/<file>/<sheet>.tsv

For every xlsx file in the locale folder (Translation.xlsx, Glossary.xlsx,
Texture.xlsx, etc.), every sheet is exported as one TSV.

Behavior:
    - All sheets exported (including MetaData)
    - Rows where every cell is None are skipped (truly empty visual separators)
    - All cell values are converted to str (avoids Excel type ambiguity)
    - Atomic write per file (.tmp + rename)
    - Stale TSVs (sheets that no longer exist in xlsx) are removed from output
    - Excel temp files ("~$..." lock files) are skipped
    - `_sheet_order.txt` is written alongside the per-sheet TSVs, listing
      the sheets in their original xlsx order (one name per line). This
      lets downstream tools (e.g. TSV -> xlsx rebuild) preserve sheet order.
    - Column widths are NOT written per-locale; the unified policy lives
      in `tools/width.json` keyed by category (MetaData / Translation /
      Glossary / Texture). Any pre-existing `_column_widths.json` here
      is treated as stale and removed.

Usage:
    python tools/utils/build_translation_tsv.py             # all locales
    python tools/utils/build_translation_tsv.py Korean      # one locale only
"""
import csv
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install -r tools/requirements.txt", file=sys.stderr)
    sys.exit(1)

# Windows console UTF-8 setup
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass


SKIP_FILE_PREFIXES = ("~$",)  # Excel lock/temp files


def _normalize_cell(value) -> str:
    """Convert an Excel cell value to its canonical string form."""
    if value is None:
        return ""
    return str(value)


def _is_empty_row(row: tuple) -> bool:
    """True if every cell in the row is None (truly empty visual row)."""
    return all(v is None for v in row)


def export_xlsx(xlsx_path: Path, out_dir: Path) -> tuple[int, int]:
    """Export each sheet of an xlsx as <out_dir>/<sheet>.tsv.

    Returns (sheet_count, total_data_rows).
    """
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        expected_tsvs: set[str] = set()
        sheet_count = 0
        total_rows = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            tsv_name = f"{sheet_name}.tsv"
            expected_tsvs.add(tsv_name)
            tsv_path = out_dir / tsv_name
            tmp = tsv_path.with_suffix(tsv_path.suffix + ".tmp")

            try:
                with open(tmp, "w", encoding="utf-8", newline="") as f:
                    writer = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)
                    sheet_rows = 0
                    for row_values in ws.iter_rows(values_only=True):
                        if row_values is None:
                            continue
                        if _is_empty_row(row_values):
                            continue
                        cells = [_normalize_cell(v) for v in row_values]
                        writer.writerow(cells)
                        sheet_rows += 1
                tmp.replace(tsv_path)
                sheet_count += 1
                total_rows += sheet_rows
                print(f"    -> {tsv_name}  ({sheet_rows} rows)")
            except Exception:
                if tmp.exists():
                    try:
                        tmp.unlink()
                    except OSError:
                        pass
                raise

        # remove stale TSVs (sheets that no longer exist in this xlsx)
        for stale in out_dir.glob("*.tsv"):
            if stale.name not in expected_tsvs:
                try:
                    stale.unlink()
                    print(f"    x removed stale: {stale.name}")
                except OSError as e:
                    print(f"    [WARN] Failed to remove stale: {stale.name} ({e})")

        # write sheet-order metadata so TSV -> xlsx rebuild preserves order
        order_path = out_dir / "_sheet_order.txt"
        order_tmp = order_path.with_suffix(order_path.suffix + ".tmp")
        try:
            with open(order_tmp, "w", encoding="utf-8", newline="\n") as f:
                for name in wb.sheetnames:
                    f.write(name + "\n")
            order_tmp.replace(order_path)
            print(f"    -> _sheet_order.txt ({len(wb.sheetnames)} sheets)")
        except Exception:
            if order_tmp.exists():
                try:
                    order_tmp.unlink()
                except OSError:
                    pass
            raise

        # column-width policy is unified in tools/width.json (not per-locale).
        # Any pre-existing _column_widths.json here is stale.
        stale_widths = out_dir / "_column_widths.json"
        if stale_widths.exists():
            try:
                stale_widths.unlink()
                print(f"    x removed stale: _column_widths.json")
            except OSError as e:
                print(f"    [WARN] Failed to remove stale: _column_widths.json ({e})")

        return sheet_count, total_rows
    finally:
        wb.close()


def process_locale(input_root: Path, locale: str,
                   output_root: Path | None = None) -> tuple[int, int, int]:
    """Process all xlsx files in one locale folder.

    Reads xlsx from input_root/<locale>/*.xlsx.
    Writes TSV to output_root/<locale>/<file_stem>/*.tsv (defaults to input_root).

    Returns (xlsx_count, sheet_count, row_count).
    """
    if output_root is None:
        output_root = input_root

    locale_dir = input_root / locale
    out_locale_dir = output_root / locale

    if not locale_dir.exists():
        print(f"[ERROR] Locale folder not found: {locale_dir}")
        return 0, 0, 0

    xlsx_files = sorted(
        f for f in locale_dir.glob("*.xlsx")
        if not f.name.startswith(SKIP_FILE_PREFIXES)
    )

    if not xlsx_files:
        print(f"[INFO] No xlsx files found in: {locale_dir}")
        return 0, 0, 0

    xlsx_count = 0
    total_sheets = 0
    total_rows = 0

    for xlsx_path in xlsx_files:
        out_dir = out_locale_dir / xlsx_path.stem
        try:
            rel_out = out_dir.relative_to(input_root.parent)
        except ValueError:
            rel_out = out_dir
        print(f"  [{locale}] {xlsx_path.name}  ->  {rel_out}/")
        try:
            sc, rc = export_xlsx(xlsx_path, out_dir)
            xlsx_count += 1
            total_sheets += sc
            total_rows += rc
        except PermissionError:
            print(f"    [ERROR] Cannot read xlsx (file open in Excel?): {xlsx_path}")
        except Exception as e:
            print(f"    [ERROR] Failed to export: {e}")

    return xlsx_count, total_sheets, total_rows


def discover_locales(translations_root: Path) -> list[str]:
    """Find locale folders that contain at least one xlsx (excluding lock files)."""
    locales = []
    for d in sorted(translations_root.iterdir()):
        if not d.is_dir():
            continue
        has_xlsx = any(
            not f.name.startswith(SKIP_FILE_PREFIXES)
            for f in d.glob("*.xlsx")
        )
        if has_xlsx:
            locales.append(d.name)
    return locales


def main() -> int:
    script_dir = Path(__file__).resolve().parent
    project_root = script_dir.parent.parent
    input_root = project_root / "Translations"

    if not input_root.exists():
        print(f"[ERROR] Translations root not found: {input_root}")
        return 1

    # Parse CLI: positional locale + optional --output-root <path>
    positional = []
    output_root = None
    rest = list(sys.argv[1:])
    i = 0
    while i < len(rest):
        a = rest[i]
        if a == "--output-root":
            if i + 1 >= len(rest):
                print("[ERROR] --output-root requires a value", file=sys.stderr)
                return 1
            output_root = Path(rest[i + 1]).resolve()
            i += 2
        elif a.startswith("--"):
            print(f"[ERROR] Unknown flag: {a}", file=sys.stderr)
            return 1
        else:
            positional.append(a)
            i += 1

    if positional:
        locales = [positional[0]]
    else:
        locales = discover_locales(input_root)

    if not locales:
        print(f"[ERROR] No locales with xlsx files found under: {input_root}")
        return 1

    effective_output = output_root if output_root else input_root
    print(f"Input   : {input_root}")
    print(f"Output  : {effective_output}")
    print(f"Locales : {', '.join(locales)}")
    print()

    total_xlsx = 0
    total_sheets = 0
    total_rows = 0

    for locale in locales:
        xc, sc, rc = process_locale(input_root, locale, output_root=output_root)
        total_xlsx += xc
        total_sheets += sc
        total_rows += rc

    print()
    print("=" * 60)
    print(f"Done: {total_xlsx} xlsx files, {total_sheets} sheets, {total_rows} rows")
    return 0


if __name__ == "__main__":
    sys.exit(main())
