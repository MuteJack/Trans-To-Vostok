"""Extract texture rework credits from <locale>/Texture.xlsx -> credits.json.

Reads the `Reworked by` and `Contributors` columns from each sheet of
`Translations/<locale>/Texture.xlsx` (MetaData sheet skipped), dedupes,
and writes the unified list to `Trans To Vostok/<locale>/credits.json`'s
`Texture_reworker` field.

Other credits.json fields (Translator, translation_updated) are preserved.

Texture.xlsx is the authoring source for image credits because texture
work happens locally with image editing tools, off-Crowdin. This script
replaces the texture-side of the old `build_authors.py` direct-extraction:
data flows xlsx -> credits.json -> AUTHORS.md (refactored later).

Usage:
    python tools/utils/get_texture_credits.py             # all active locales
    python tools/utils/get_texture_credits.py Korean      # one locale
"""
import argparse
import json
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install -r tools/requirements.txt",
          file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = Path(__file__).resolve().parent
TOOLS_DIR = SCRIPT_DIR.parent
REPO = TOOLS_DIR.parent

sys.path.insert(0, str(TOOLS_DIR))
from utils.locale_config import load_crowdin_locale_map

MOD_ROOT = REPO / "Trans To Vostok"
TRANSLATIONS_ROOT = REPO / "Translations"
CREDITS_FILE = "credits.json"
TEXTURE_XLSX = "Texture.xlsx"
TEXTURE_REWORKED_COLUMN = "Reworked by"
TEXTURE_CONTRIBUTORS_COLUMN = "Contributors"


def _split_names(value) -> list[str]:
    """Split a cell value into trimmed names. Excel allows in-cell newlines (Alt+Enter)."""
    if value is None:
        return []
    if not isinstance(value, str):
        s = str(value).strip()
        return [s] if s else []
    return [n.strip() for n in value.replace("\r\n", "\n").split("\n") if n.strip()]


def collect_texture_credits(xlsx_path: Path) -> list[str]:
    """Ordered, deduped list of names from `Reworked by` + `Contributors` across all sheets."""
    if not xlsx_path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"  [WARN] Cannot read {xlsx_path}: {e}", file=sys.stderr)
        return []

    seen: set[str] = set()
    out: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c).strip() if c is not None else "" for c in rows[0]]
            rb_idx = header.index(TEXTURE_REWORKED_COLUMN) if TEXTURE_REWORKED_COLUMN in header else None
            co_idx = header.index(TEXTURE_CONTRIBUTORS_COLUMN) if TEXTURE_CONTRIBUTORS_COLUMN in header else None
            for row in rows[1:]:
                if row is None:
                    continue
                for idx in (rb_idx, co_idx):
                    if idx is None or idx >= len(row):
                        continue
                    for name in _split_names(row[idx]):
                        if name not in seen:
                            seen.add(name)
                            out.append(name)
    finally:
        wb.close()
    return out


def update_credits_for_locale(locale: str) -> list[str]:
    """Read Texture.xlsx, write Texture_reworker field. Preserves other fields."""
    xlsx = TRANSLATIONS_ROOT / locale / TEXTURE_XLSX
    names = collect_texture_credits(xlsx)

    credits_path = MOD_ROOT / locale / CREDITS_FILE
    existing: dict = {}
    if credits_path.exists():
        try:
            existing = json.loads(credits_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            print(f"  [WARN] {credits_path} is not valid JSON; recreating from scratch.")
            existing = {}

    existing["Texture_reworker"] = names

    credits_path.parent.mkdir(parents=True, exist_ok=True)
    credits_path.write_text(
        json.dumps(existing, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    return names


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "locale", nargs="?", default=None,
        help="Locale folder name (one of active). Omit to process all active locales.",
    )
    args = parser.parse_args(argv[1:])

    locale_map = load_crowdin_locale_map()
    if args.locale is not None:
        if args.locale not in locale_map:
            print(f"[ERROR] Unknown locale: {args.locale}", file=sys.stderr)
            print(f"        Known: {', '.join(locale_map)}", file=sys.stderr)
            return 1
        locales = [args.locale]
    else:
        locales = list(locale_map)

    print(f"Locales: {', '.join(locales)}")
    print()

    for loc in locales:
        print(f"=== {loc} ===")
        names = update_credits_for_locale(loc)
        if names:
            print(f"  Texture_reworker: {len(names)} ({', '.join(names)})")
        else:
            print(f"  Texture_reworker: 0 (none)")
        print()

    print("Done.")
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
