"""Build <pkg_root>/info.json — metadata consumed by F9 UI's Info tab.

Sources:
  - mod_version           : mod.txt          (`version="..."`)
  - build_date            : today's date (UTC)
  - target_game_version   : <Korean or first-found locale>/Translation.xlsx
                            MetaData → "Game Version"
  - lead_developer / code_contributors / acknowledgments
                          : AUTHORS.md (sections under `## ...`)
  - locales.<locale>.translation_updated
                          : Translation.xlsx MetaData → "Translation Updated Date"
  - locales.<locale>.texture_updated
                          : (currently unavailable — Texture.xlsx has no
                             MetaData sheet; reports "(unavailable)")
  - locales.<locale>.translators / translation_contributors
                          : Translation.xlsx MetaData (line-break separated)
  - locales.<locale>.texture_reworkers / texture_contributors
                          : Texture.xlsx data sheets — union of "Reworked by"
                             and "Contributors" columns

Output:
  <pkg_root>/info.json (UTF-8, indent=2)

This script is best-effort: per-locale read failures or missing fields
are reported as warnings but do not fail the script. Defaults ensure the
JSON always has the expected keys.

Usage:
  python tools/utils/build_mod_info.py
"""
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. pip install openpyxl", file=sys.stderr)
    sys.exit(1)


METADATA_SHEET = "MetaData"
TEXTURE_REWORKER_COL = "Reworked by"
TEXTURE_CONTRIB_COL = "Contributors"


def parse_mod_version(mod_txt: Path) -> str:
    if not mod_txt.exists():
        return "unknown"
    try:
        text = mod_txt.read_text(encoding="utf-8")
    except OSError:
        return "unknown"
    for line in text.splitlines():
        m = re.match(r'\s*version\s*=\s*"?([^"\n]+?)"?\s*$', line)
        if m:
            return m.group(1).strip()
    return "unknown"


def parse_authors_by_section(authors_md: Path) -> dict:
    """AUTHORS.md → { 'lead_developer': [...], 'code_contributors': [...],
    'acknowledgments': [...] }. Reads `- **Name**` patterns under each `##`.
    Skips the auto-generated Translators section (handled separately)."""
    out = {"lead_developer": [], "code_contributors": [], "acknowledgments": []}
    if not authors_md.exists():
        return out
    try:
        text = authors_md.read_text(encoding="utf-8")
    except OSError:
        return out

    section_map = {
        "Author / Lead Developer": "lead_developer",
        "Code Contributors": "code_contributors",
        "Acknowledgments": "acknowledgments",
    }
    current_key = None
    for line in text.splitlines():
        h = re.match(r'##\s+(.+?)\s*$', line)
        if h:
            current_key = section_map.get(h.group(1).strip())
            continue
        if current_key is None:
            continue
        n = re.match(r'-\s*\*\*([^*]+)\*\*', line)
        if not n:
            continue
        name = n.group(1).strip()
        if not name or name.lower() in {"none", "unknown", "tbd"}:
            continue
        if name not in out[current_key]:
            out[current_key].append(name)
    return out


def _split_names(s: str) -> list[str]:
    """Cell value may have multiple names separated by line breaks."""
    if not s:
        return []
    return [n.strip() for n in str(s).splitlines() if n.strip()]


def read_translation_metadata(xlsx_path: Path) -> dict:
    """Translation.xlsx MetaData → flat dict of Field→Value. {} on any error."""
    if not xlsx_path.exists():
        return {}
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"[WARN] could not open {xlsx_path}: {e}", file=sys.stderr)
        return {}
    meta: dict[str, str] = {}
    try:
        if METADATA_SHEET not in wb.sheetnames:
            return {}
        ws = wb[METADATA_SHEET]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return {}
        # header expected: IDX, Field, Value, Description
        h = [str(c) if c is not None else "" for c in rows[0]]
        try:
            f_i = h.index("Field")
            v_i = h.index("Value")
        except ValueError:
            return {}
        for r in rows[1:]:
            if r is None or len(r) <= max(f_i, v_i):
                continue
            field = r[f_i]
            value = r[v_i]
            if field is None:
                continue
            meta[str(field).strip()] = "" if value is None else str(value)
    finally:
        wb.close()
    return meta


def read_texture_credits(xlsx_path: Path) -> dict:
    """Texture.xlsx → {'reworkers': sorted unique, 'contributors': sorted unique}."""
    out = {"reworkers": [], "contributors": []}
    if not xlsx_path.exists():
        return out
    try:
        wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"[WARN] could not open {xlsx_path}: {e}", file=sys.stderr)
        return out
    rew_set: set[str] = set()
    con_set: set[str] = set()
    try:
        for sname in wb.sheetnames:
            if sname == METADATA_SHEET:
                continue
            ws = wb[sname]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c) if c is not None else "" for c in rows[0]]
            try:
                rew_i = header.index(TEXTURE_REWORKER_COL)
            except ValueError:
                rew_i = -1
            try:
                con_i = header.index(TEXTURE_CONTRIB_COL)
            except ValueError:
                con_i = -1
            for r in rows[1:]:
                if r is None:
                    continue
                if rew_i >= 0 and rew_i < len(r):
                    for n in _split_names(r[rew_i]):
                        rew_set.add(n)
                if con_i >= 0 and con_i < len(r):
                    for n in _split_names(r[con_i]):
                        con_set.add(n)
    finally:
        wb.close()
    out["reworkers"] = sorted(rew_set)
    out["contributors"] = sorted(con_set)
    return out


def collect_locale(translations_root: Path, locale: str) -> dict:
    info = {
        "translation_updated": "unknown",
        "texture_updated": "unknown",
        "translators": [],
        "translation_contributors": [],
        "texture_reworkers": [],
        "texture_contributors": [],
    }
    # Translation.xlsx MetaData
    trans_meta = read_translation_metadata(translations_root / locale / "Translation.xlsx")
    if trans_meta:
        date_str = trans_meta.get("Translation Updated Date", "").strip()
        if date_str:
            info["translation_updated"] = date_str.split(" ", 1)[0]  # date only
        info["translators"] = _split_names(trans_meta.get("Translator", ""))
        info["translation_contributors"] = _split_names(trans_meta.get("Contributor (Translate)", ""))

    # Texture.xlsx — first try MetaData sheet, fall back to Translation date.
    tex_path = translations_root / locale / "Texture.xlsx"
    tex_meta = read_translation_metadata(tex_path)  # same shape (Field/Value)
    tex_date = tex_meta.get("Texture Updated Date", "").strip() if tex_meta else ""
    if tex_date:
        info["texture_updated"] = tex_date.split(" ", 1)[0]
    elif info["translation_updated"] != "unknown":
        # fallback: textures are usually updated alongside translations
        info["texture_updated"] = info["translation_updated"]

    tex = read_texture_credits(tex_path)
    info["texture_reworkers"] = tex["reworkers"]
    info["texture_contributors"] = tex["contributors"]
    return info


def discover_locales(translations_root: Path) -> list[str]:
    """List subdirs that look like locale folders (contain Translation.xlsx)."""
    if not translations_root.exists():
        return []
    out = []
    for d in sorted(translations_root.iterdir()):
        if d.is_dir() and (d / "Translation.xlsx").exists():
            out.append(d.name)
    return out


def derive_target_game_version(translations_root: Path, locales: list[str]) -> str:
    """Use Korean if available; otherwise the first locale with a value."""
    preferred = ["Korean"] + [l for l in locales if l != "Korean"]
    for loc in preferred:
        meta = read_translation_metadata(translations_root / loc / "Translation.xlsx")
        v = meta.get("Game Version", "").strip()
        if v:
            return v
    return "unknown"


def build_info(repo_root: Path) -> dict:
    translations_root = repo_root / "Translations"
    locales = [l for l in discover_locales(translations_root) if l != "Template"]
    sections = parse_authors_by_section(repo_root / "AUTHORS.md")
    info = {
        "mod_version": parse_mod_version(repo_root / "mod.txt"),
        "build_date": datetime.now(timezone.utc).strftime("%Y-%m-%d"),
        "target_game_version": derive_target_game_version(translations_root, locales),
        "lead_developer": sections["lead_developer"],
        "code_contributors": sections["code_contributors"],
        "acknowledgments": sections["acknowledgments"],
        "locales": {locale: collect_locale(translations_root, locale) for locale in locales},
    }
    return info


def main() -> int:
    script_dir = Path(__file__).resolve().parent
    repo_root = script_dir.parent.parent
    pkg_root = repo_root / "Trans To Vostok"
    output_path = pkg_root / "info.json"

    info = build_info(repo_root)
    try:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        output_path.write_text(
            json.dumps(info, indent=2, ensure_ascii=False),
            encoding="utf-8",
        )
    except OSError as e:
        print(f"[ERROR] Failed to write {output_path}: {e}", file=sys.stderr)
        return 1
    print(f"[OK] Wrote {output_path.relative_to(repo_root)}")
    print(f"  mod_version={info['mod_version']}, build_date={info['build_date']}")
    print(f"  target_game_version={info['target_game_version']}")
    print(f"  lead_developer={len(info['lead_developer'])}, "
          f"code_contributors={len(info['code_contributors'])}, "
          f"acknowledgments={len(info['acknowledgments'])}")
    print(f"  locales: {list(info['locales'].keys())}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
