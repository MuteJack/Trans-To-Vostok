"""
Auto-update the Translators section of AUTHORS.md.

Reads each locale's:
  - Translation.xlsx MetaData ("Translator", "Contributor (Translate)")
  - Texture.xlsx ("Reworked by", "Contributors" columns)

and aggregates them into a single Translators section in the project-root
AUTHORS.md, between BEGIN/END markers. Manual sections (Author / Lead
Developer, Code Contributors, How to add yourself) are preserved.

Markers (must already exist in AUTHORS.md):
    <!-- BEGIN AUTO-GENERATED: Translators -->
    ...replaced content...
    <!-- END AUTO-GENERATED: Translators -->

Usage:
    python tools/utils/build_authors.py
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install -r tools/requirements.txt", file=sys.stderr)
    sys.exit(1)

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass


METADATA_TRANSLATOR_FIELD = "Translator"
METADATA_CONTRIBUTOR_FIELD = "Contributor (Translate)"
TEXTURE_REWORKED_COLUMN = "Reworked by"
TEXTURE_CONTRIBUTORS_COLUMN = "Contributors"

BEGIN_MARKER = "<!-- BEGIN AUTO-GENERATED: Translators -->"
END_MARKER = "<!-- END AUTO-GENERATED: Translators -->"


def _split_names(value) -> list[str]:
    if value is None:
        return []
    return [line.strip() for line in str(value).split("\n") if line.strip()]


def _read_metadata_field(xlsx_path: Path, field_name: str) -> list[str]:
    if not xlsx_path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except (PermissionError, Exception) as e:
        print(f"  [WARN] Cannot read {xlsx_path.name}: {e}")
        return []
    try:
        if "MetaData" not in wb.sheetnames:
            return []
        ws = wb["MetaData"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        header = [str(c).strip() if c is not None else "" for c in rows[0]]
        try:
            field_idx = header.index("Field")
            value_idx = header.index("Value")
        except ValueError:
            return []
        for row in rows[1:]:
            if row is None or len(row) <= max(field_idx, value_idx):
                continue
            field = str(row[field_idx]).strip() if row[field_idx] is not None else ""
            if field == field_name:
                return _split_names(row[value_idx])
        return []
    finally:
        wb.close()


def _collect_texture_credits(xlsx_path: Path) -> tuple[set[str], set[str]]:
    reworked: set[str] = set()
    contributors: set[str] = set()
    if not xlsx_path.exists():
        return reworked, contributors
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except (PermissionError, Exception) as e:
        print(f"  [WARN] Cannot read {xlsx_path.name}: {e}")
        return reworked, contributors
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name == "MetaData":
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c).strip() if c is not None else "" for c in rows[0]]
            rb_idx = header.index(TEXTURE_REWORKED_COLUMN) if TEXTURE_REWORKED_COLUMN in header else None
            co_idx = header.index(TEXTURE_CONTRIBUTORS_COLUMN) if TEXTURE_CONTRIBUTORS_COLUMN in header else None
            for row in rows[1:]:
                if row is None:
                    continue
                if rb_idx is not None and rb_idx < len(row):
                    for n in _split_names(row[rb_idx]):
                        reworked.add(n)
                if co_idx is not None and co_idx < len(row):
                    for n in _split_names(row[co_idx]):
                        contributors.add(n)
    finally:
        wb.close()
    return reworked, contributors


def _discover_locales(translations_root: Path) -> list[str]:
    locales: list[str] = []
    for d in sorted(translations_root.iterdir()):
        if not d.is_dir():
            continue
        if (d / "Translation.xlsx").exists() or (d / "Texture.xlsx").exists():
            locales.append(d.name)
    return locales


def _render_subsection(label: str, names: list[str]) -> list[str]:
    if names:
        out = [f"**{label}:**"]
        for n in names:
            out.append(f"- {n}")
        out.append("")
    else:
        out = [f"**{label}:** _(none yet)_", ""]
    return out


def _render_locale(locale: str, lead: list[str], translation_contrib: list[str],
                   texture_reworked: list[str], texture_contrib: list[str]) -> list[str]:
    lines: list[str] = []
    lines.append(f"### {locale} (`Translations/{locale}/`)")
    lines.append("")
    lines.extend(_render_subsection("Lead Translator(s)", lead))
    lines.extend(_render_subsection("Translation Contributors", translation_contrib))
    lines.extend(_render_subsection("Texture Reworkers", texture_reworked))
    lines.extend(_render_subsection("Texture Contributors", texture_contrib))
    return lines


def build_auto_section(translations_root: Path) -> str:
    """Build the markdown content for the auto-generated Translators section."""
    locales = _discover_locales(translations_root)

    lines: list[str] = []
    lines.append("## Translators")
    lines.append("")
    lines.append(
        "_Auto-generated from each locale's `Translation.xlsx` MetaData "
        "(`Translator`, `Contributor (Translate)`) and `Texture.xlsx` "
        "(`Reworked by`, `Contributors`) by `tools/utils/build_authors.py`. "
        "Update the source xlsx files to change this list._"
    )
    lines.append("")

    if not locales:
        lines.append("_(no locales registered yet)_")
        return "\n".join(lines) + "\n"

    for locale in locales:
        locale_dir = translations_root / locale
        translation_xlsx = locale_dir / "Translation.xlsx"
        texture_xlsx = locale_dir / "Texture.xlsx"

        lead = _read_metadata_field(translation_xlsx, METADATA_TRANSLATOR_FIELD)
        translation_contrib = _read_metadata_field(translation_xlsx, METADATA_CONTRIBUTOR_FIELD)
        texture_reworked_set, texture_contrib_set = _collect_texture_credits(texture_xlsx)

        # dedup within MetaData (Translator vs Contributor mutually exclusive)
        seen_meta: set[str] = set()
        lead_unique: list[str] = []
        for n in lead:
            if n not in seen_meta:
                lead_unique.append(n)
                seen_meta.add(n)
        translation_contrib_unique: list[str] = []
        for n in translation_contrib:
            if n not in seen_meta:
                translation_contrib_unique.append(n)
                seen_meta.add(n)

        # texture sections are independent of MetaData (same person can be in both)
        texture_reworked = sorted(texture_reworked_set)
        texture_contrib = sorted(texture_contrib_set - texture_reworked_set)

        lines.extend(_render_locale(
            locale, lead_unique, translation_contrib_unique,
            texture_reworked, texture_contrib,
        ))

    return "\n".join(lines) + "\n"


def update_authors_md(authors_path: Path, generated: str) -> bool:
    """Replace content between BEGIN/END markers. Returns True on success."""
    if not authors_path.exists():
        print(f"[ERROR] AUTHORS.md not found: {authors_path}")
        return False

    text = authors_path.read_text(encoding="utf-8")
    if BEGIN_MARKER not in text or END_MARKER not in text:
        print(f"[ERROR] AUTHORS.md is missing one or both markers:")
        print(f"    {BEGIN_MARKER}")
        print(f"    {END_MARKER}")
        print(f"  Add these markers around the section to be auto-generated.")
        return False

    before, _, rest = text.partition(BEGIN_MARKER)
    _, _, after = rest.partition(END_MARKER)

    new_text = f"{before}{BEGIN_MARKER}\n\n{generated}\n{END_MARKER}{after}"

    if new_text == text:
        print(f"  -> No changes (already up-to-date): {authors_path.name}")
        return True

    tmp = authors_path.with_suffix(authors_path.suffix + ".tmp")
    try:
        tmp.write_text(new_text, encoding="utf-8")
        tmp.replace(authors_path)
    except Exception:
        if tmp.exists():
            try:
                tmp.unlink()
            except OSError:
                pass
        raise
    print(f"  -> Updated: {authors_path.name}")
    return True


def main() -> int:
    script_dir = Path(__file__).resolve().parent
    project_root = script_dir.parent.parent
    translations_root = project_root / "Translations"
    authors_path = project_root / "AUTHORS.md"

    if not translations_root.exists():
        print(f"[ERROR] Translations root not found: {translations_root}")
        return 1
    if not authors_path.exists():
        print(f"[ERROR] AUTHORS.md not found: {authors_path}")
        return 1

    print(f"Project root: {project_root}")
    print(f"Sources     : {translations_root}")
    print(f"Target      : {authors_path}")
    print()

    locales = _discover_locales(translations_root)
    print(f"Locales found: {', '.join(locales) if locales else '(none)'}")
    print()

    auto_content = build_auto_section(translations_root)
    if not update_authors_md(authors_path, auto_content):
        return 1
    return 0


if __name__ == "__main__":
    sys.exit(main())
