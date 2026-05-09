"""
Import DeepL translation results into a target locale's xlsx files.

Reads:
    .tmp/unique_text/<source>/translated_<DEEPL_LANG>.tsv   (DeepL results)
Writes (in-place, all under Trans To Vostok/<target_locale>/):
    Translation.xlsx   (game text)
    Texture.xlsx       (image labels, capitalized columns)
    Glossary.xlsx      (translator reference)

Per-row logic (per file, columns vary; checks adapted by config):
    1. If row's translation is already non-empty:
         skip (preserve human edits / curated entries / previous runs)
    2. If untranslatable == "1" (only when column exists):
         translation = text (copy original)
         Comments NOT updated (it's a copy, not a machine translation)
    3. Else if method == "ignore" (only when column exists):
         look up text in translated TSV (text-based)
         if found, use that translation; append `#Machine Translated` to Comments
         if not found, fallback: copy text to translation
         (an ignore row is usually a duplicate of a non-ignore row whose text
          was translated; reusing that translation keeps the xlsx complete)
    4. Else if method == "pattern" (only when column exists):
         skip (DeepL cannot translate regex source patterns)
    5. Else (regular: static / literal / substr / Texture / Glossary):
         look up text in translated TSV
         if found, write translation; append `#Machine Translated` to Comments

Note: lookup is by exact text match on the row's text column. The unique.tsv
/ mapping.tsv files are not needed here — translated_<TARGET>.tsv already
contains (text, translation) pairs.

Usage:
    python tools/utils/import_translations.py <target_locale> [--deepl-lang <code>] [--source <locale>]

Examples:
    python tools/utils/import_translations.py French --deepl-lang FR
    python tools/utils/import_translations.py Japanese --deepl-lang JA
    python tools/utils/import_translations.py French                 # deepl-lang defaults to FR
"""
import argparse
import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install -r tools/requirements.txt", file=sys.stderr)
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
from utils.locale_config import dir_to_deepl_id, default_source_locale

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass


BOOL_TRUE = {"1", "true"}
MT_MARKER = "#Machine Translated"


def append_marker(existing: str, marker: str) -> str:
    existing = existing or ""
    if marker in existing:
        return existing
    if existing.strip():
        return existing + "\n" + marker
    return marker

# Per-xlsx column configuration (None = column not present in this file)
XLSX_FILES = [
    {
        "name": "Translation.xlsx",
        "text_col": "text",
        "trans_col": "translation",
        "method_col": "method",
        "untrans_col": "untranslatable",
        "comments_col": "Comments",
    },
    {
        "name": "Texture.xlsx",
        "text_col": "Text",
        "trans_col": "Translation",
        "method_col": None,
        "untrans_col": None,
        "comments_col": "Comments",
    },
    {
        "name": "Glossary.xlsx",
        "text_col": "text",
        "trans_col": "translation",
        "method_col": None,
        "untrans_col": "untranslatable",
        "comments_col": "Comments",
    },
]


def load_translated_map(translated_path: Path) -> dict[str, dict]:
    """Build text -> {translation, status, unique_id} map from translated TSV."""
    if not translated_path.exists():
        print(f"[ERROR] Translated TSV not found: {translated_path}", file=sys.stderr)
        return {}
    out: dict[str, dict] = {}
    with open(translated_path, "r", encoding="utf-8", newline="") as f:
        reader = csv.DictReader(f, delimiter="\t")
        for row in reader:
            text = row.get("source", "")
            translation = row.get("translation", "")
            status = row.get("status", "")
            uid = row.get("unique_id", "")
            # only "ok" or "placeholder_lost" rows have usable translations
            if not translation or status == "error":
                continue
            # if duplicate text appears (shouldn't, since unique.tsv was deduped),
            # keep the first one
            if text in out:
                continue
            out[text] = {
                "translation": translation,
                "status": status,
                "unique_id": uid,
            }
    return out


def _new_stats() -> dict:
    return {
        "rows_total": 0,
        "rows_skipped_already_translated": 0,
        "rows_filled_untranslatable": 0,
        "rows_filled_ignore": 0,
        "rows_filled_ignore_fallback": 0,
        "rows_filled_regular": 0,
        "rows_skipped_pattern": 0,
        "rows_no_text": 0,
        "rows_no_translation_available": 0,
        "rows_placeholder_warning": 0,
    }


def import_to_xlsx(target_xlsx: Path, cfg: dict, translated_map: dict[str, dict],
                   stats: dict) -> None:
    """Apply translations to one xlsx using its column config. Updates stats in place."""
    if not target_xlsx.exists():
        return

    wb = openpyxl.load_workbook(target_xlsx)
    try:
        for sheet_name in wb.sheetnames:
            if sheet_name == "MetaData":
                continue
            ws = wb[sheet_name]
            if ws.max_row < 2:
                continue

            header_row = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
            header = {
                str(h).strip(): i + 1
                for i, h in enumerate(header_row) if h is not None
            }

            text_col = header.get(cfg["text_col"])
            trans_col = header.get(cfg["trans_col"])
            if text_col is None or trans_col is None:
                continue  # not a translation sheet

            method_col = header.get(cfg["method_col"]) if cfg.get("method_col") else None
            untrans_col = header.get(cfg["untrans_col"]) if cfg.get("untrans_col") else None
            comments_col = header.get(cfg["comments_col"]) if cfg.get("comments_col") else None

            for row_idx in range(2, ws.max_row + 1):
                stats["rows_total"] += 1

                text = ws.cell(row_idx, text_col).value
                if text is None or str(text).strip() == "":
                    stats["rows_no_text"] += 1
                    continue
                text = str(text)

                existing_translation = ws.cell(row_idx, trans_col).value
                if existing_translation is not None and str(existing_translation).strip() != "":
                    stats["rows_skipped_already_translated"] += 1
                    continue

                untrans_val = ""
                if untrans_col is not None:
                    v = ws.cell(row_idx, untrans_col).value
                    untrans_val = str(v).strip().lower() if v is not None else ""
                method_val = ""
                if method_col is not None:
                    v = ws.cell(row_idx, method_col).value
                    method_val = str(v).strip().lower() if v is not None else ""

                # 2. untranslatable → copy original; MT NOT set
                if untrans_val in BOOL_TRUE:
                    ws.cell(row_idx, trans_col).value = text
                    stats["rows_filled_untranslatable"] += 1
                    continue

                # 4. pattern → skip
                if method_val == "pattern":
                    stats["rows_skipped_pattern"] += 1
                    continue

                # 3 & 5. ignore or regular → look up text in translated map
                entry = translated_map.get(text)
                if entry is None:
                    if method_val == "ignore":
                        ws.cell(row_idx, trans_col).value = text
                        stats["rows_filled_ignore_fallback"] += 1
                    else:
                        stats["rows_no_translation_available"] += 1
                    continue

                ws.cell(row_idx, trans_col).value = entry["translation"]
                if comments_col is not None:
                    existing = ws.cell(row_idx, comments_col).value
                    new_val = append_marker(str(existing) if existing is not None else "", MT_MARKER)
                    ws.cell(row_idx, comments_col).value = new_val
                if entry["status"] == "placeholder_lost":
                    stats["rows_placeholder_warning"] += 1

                if method_val == "ignore":
                    stats["rows_filled_ignore"] += 1
                else:
                    stats["rows_filled_regular"] += 1

        wb.save(target_xlsx)
    finally:
        wb.close()


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Import DeepL translation results into target locale's Translation.xlsx"
    )
    parser.add_argument("target_locale", help="Target locale folder name (e.g., French, Japanese)")
    parser.add_argument(
        "--deepl-lang", default=None,
        help="DeepL language code (e.g., FR, JA). Defaults based on target_locale."
    )
    parser.add_argument(
        "--source", default=None,
        help="Source locale folder where unique.tsv lives (default: same as target_locale)"
    )
    args = parser.parse_args()

    target_locale = args.target_locale
    source_locale = args.source or target_locale

    project_source = default_source_locale()
    if target_locale == project_source:
        print(
            f"[ERROR] '{target_locale}' is the project's source language "
            f"(declared in tools/languages.json:default_source).\n"
            f"  Imports go INTO a target locale — pick a target.",
            file=sys.stderr,
        )
        return 1

    deepl_lang = args.deepl_lang or dir_to_deepl_id(target_locale)
    if not deepl_lang:
        print(
            f"[ERROR] Cannot determine DeepL language code for locale '{target_locale}'.\n"
            f"  Specify it explicitly with --deepl-lang.",
            file=sys.stderr,
        )
        return 1

    script_dir = Path(__file__).resolve().parent
    mod_root = script_dir.parent.parent
    translations_root = mod_root / "Translations"

    locale_dir = translations_root / target_locale
    translated_path = mod_root / ".tmp" / "unique_text" / source_locale / f"translated_{deepl_lang}.tsv"

    if not locale_dir.exists():
        print(f"[ERROR] Target locale folder not found: {locale_dir}", file=sys.stderr)
        return 1
    if not translated_path.exists():
        print(f"[ERROR] Translated TSV not found: {translated_path}", file=sys.stderr)
        print(f"  Run: python tools/translate_with_deepl.py {deepl_lang} --source {source_locale}", file=sys.stderr)
        return 1

    print(f"Target locale   : {target_locale}")
    print(f"Locale folder   : {locale_dir}")
    print(f"Translated TSV  : {translated_path}")
    print(f"DeepL language  : {deepl_lang}")
    print(f"Source locale   : {source_locale}")
    print()

    print("Loading translations...")
    translated_map = load_translated_map(translated_path)
    print(f"  -> {len(translated_map)} unique text-to-translation pairs")
    print()

    # Process each xlsx file with its config
    per_file_stats: dict[str, dict] = {}
    for cfg in XLSX_FILES:
        xlsx_path = locale_dir / cfg["name"]
        if not xlsx_path.exists():
            print(f"  [SKIP] {cfg['name']} not found")
            continue
        print(f"Importing into {cfg['name']}...")
        stats = _new_stats()
        try:
            import_to_xlsx(xlsx_path, cfg, translated_map, stats)
        except PermissionError:
            print(f"  [ERROR] Cannot write (file open in Excel?): {xlsx_path}", file=sys.stderr)
            return 1
        per_file_stats[cfg["name"]] = stats

    print()
    print("=" * 60)
    print(f"Done: {target_locale}/ xlsx files updated")
    for fname, stats in per_file_stats.items():
        print()
        print(f"[{fname}]")
        print(f"  Rows scanned                 : {stats['rows_total']}")
        print(f"  Skipped (already translated) : {stats['rows_skipped_already_translated']}")
        print(f"  Filled (untranslatable copy) : {stats['rows_filled_untranslatable']}")
        print(f"  Filled (ignore via lookup)   : {stats['rows_filled_ignore']}")
        print(f"  Filled (ignore fallback copy): {stats['rows_filled_ignore_fallback']}")
        print(f"  Filled (regular row)         : {stats['rows_filled_regular']}")
        print(f"  Skipped (method=pattern)     : {stats['rows_skipped_pattern']}")
        print(f"  Skipped (no text in row)     : {stats['rows_no_text']}")
        print(f"  No translation available     : {stats['rows_no_translation_available']}")
        if stats["rows_placeholder_warning"]:
            print(f"  WARNING: placeholder issues  : {stats['rows_placeholder_warning']} rows")
    return 0


if __name__ == "__main__":
    sys.exit(main())
