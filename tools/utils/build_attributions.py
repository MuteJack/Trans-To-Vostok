"""
Auto-generate Attribution.md from Images.xlsx.

Reads the following columns from each sheet and outputs as Markdown:
    - File Name
    - Reworked by
    - Attribution

Usage:
    python tools/build_attributions.py                       # default (Korean → Attribution.md)
    python tools/build_attributions.py --locale Korean
    python tools/build_attributions.py --output custom.md
"""
import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install openpyxl", file=sys.stderr)
    sys.exit(1)

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass


URL_RE = re.compile(r'https?://[^\s)\]]+')

REQUIRED_COLUMNS = ["File Name", "Reworked by", "Attribution"]


def collect_rows(xlsx_path: Path) -> list[dict]:
    """Collect (sheet, file_name, reworked_by, attribution) rows from all sheets."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    rows: list[dict] = []
    for ws in wb.worksheets:
        header = [c.value for c in ws[1]]
        try:
            idx = {col: header.index(col) for col in REQUIRED_COLUMNS}
        except ValueError as e:
            print(f"[WARN] '{ws.title}' sheet missing required column ({e}), skipping",
                  file=sys.stderr)
            continue

        for row_no, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_no == 1:
                continue
            def cell(name: str) -> str:
                i = idx[name]
                return str(row[i]).strip() if i < len(row) and row[i] is not None else ""

            file_name = cell("File Name")
            if not file_name:
                continue  # skip empty row

            rows.append({
                "sheet": ws.title,
                "file_name": file_name,
                "reworked_by": cell("Reworked by"),
                "attribution": cell("Attribution"),
            })
    return rows


def linkify(text: str) -> str:
    """Convert URLs inside text to markdown auto-links."""
    return URL_RE.sub(lambda m: f"<{m.group(0)}>", text)


def render_markdown(rows: list[dict], locale: str) -> str:
    with_attr = [r for r in rows if r["attribution"]]
    without_attr = [r for r in rows if not r["attribution"]]

    lines: list[str] = []
    lines.append(f"# Attribution — Trans To Vostok ({locale})")
    lines.append("")
    lines.append(
        "This document lists attributions for translated image assets in this mod."
    )
    lines.append("")
    lines.append(
        f"_Auto-generated from `{locale}/Images.xlsx` by "
        "`tools/build_attributions.py`. Do not edit manually — "
        "update the xlsx instead._"
    )
    lines.append("")

    if with_attr:
        lines.append("## Files with third-party attribution")
        lines.append("")
        for r in with_attr:
            lines.append(f"### `{r['file_name']}` _({r['sheet']})_")
            lines.append("")
            lines.append(f"- **Reworked by**: {r['reworked_by'] or '(unknown)'}")
            lines.append("- **Sources**:")
            for src_line in r["attribution"].splitlines():
                src_line = src_line.strip()
                if not src_line:
                    continue
                lines.append(f"  - {linkify(src_line)}")
            lines.append("")

    if without_attr:
        lines.append("## Files without third-party attribution")
        lines.append("")
        lines.append(
            "Created by the listed translator(s) without using third-party sources."
        )
        lines.append("")
        # group by sheet
        by_sheet: dict[str, list[dict]] = {}
        for r in without_attr:
            by_sheet.setdefault(r["sheet"], []).append(r)
        for sheet_name in by_sheet:
            lines.append(f"### {sheet_name}")
            lines.append("")
            for r in by_sheet[sheet_name]:
                rework = r["reworked_by"] or "(unknown)"
                lines.append(f"- `{r['file_name']}` — {rework}")
            lines.append("")

    if not rows:
        lines.append("_No image assets registered._")
        lines.append("")

    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate Attribution.md from Images.xlsx"
    )
    parser.add_argument(
        "--locale", default="Korean",
        help="Locale folder name (default: Korean)",
    )
    parser.add_argument(
        "--output", default=None,
        help="Output file path (default: <pkg_root>/<locale>/Attribution.md)",
    )
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    # script_dir = mods/Trans To Vostok/tools/utils
    mod_root = script_dir.parent.parent
    pkg_root = mod_root / "Trans To Vostok"
    xlsx_path = pkg_root / args.locale / "Images.xlsx"

    if not xlsx_path.exists():
        # When called from the build pipeline, locales without Images.xlsx are normal (no texture translation)
        print(f"[SKIP] Images.xlsx not found: {xlsx_path} - no output")
        return 0

    print(f"Input:  {xlsx_path}")
    rows = collect_rows(xlsx_path)
    sheets = sorted({r['sheet'] for r in rows})
    print(f"Loaded: {len(rows)} entries from sheets: {', '.join(sheets) or '-'}")

    md = render_markdown(rows, args.locale)

    out_path = (
        Path(args.output).resolve()
        if args.output
        else (pkg_root / args.locale / "Attribution.md")
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(md, encoding="utf-8")
    print(f"Output: {out_path}")

    with_attr = sum(1 for r in rows if r["attribution"])
    without_attr = len(rows) - with_attr
    print()
    print(f"  with attribution:    {with_attr}")
    print(f"  without attribution: {without_attr}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
