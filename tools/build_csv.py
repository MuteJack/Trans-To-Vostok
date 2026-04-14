"""
xlsx → CSV (UTF-8) 변환 스크립트

사용법:
    python build_csv.py                          # 기본 위치의 모든 xlsx 변환
    python build_csv.py path/to/file.xlsx        # 특정 파일만 변환
    python build_csv.py path/to/file.xlsx -o out.csv  # 출력 경로 지정

기본 동작:
    mods/Trans To Vostok/<locale>/*.xlsx 를 찾아서
    같은 폴더의 동일 이름 *.csv 로 내보냄.

변환 규칙:
    - 인코딩: UTF-8 (BOM 없음, Godot 기본 호환)
    - 구분자: 쉼표
    - 따옴표: 필요 시 자동 (QUOTE_MINIMAL)
    - 빈 행 (전부 None) 제거
    - Excel 아티팩트 _x000D_ 제거
    - 모든 셀은 문자열로 변환 (int/float/None 포함)
"""
import argparse
import csv
import sys
from pathlib import Path

# Windows 콘솔 한글 출력 지원
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl이 설치되어 있지 않습니다.", file=sys.stderr)
    print("설치: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def clean_cell(value):
    """셀 값을 CSV에 적합한 문자열로 변환."""
    if value is None:
        return ""
    if isinstance(value, str):
        return value.replace("_x000D_", "").rstrip()
    # int, float, bool 등
    return str(value)


def is_empty_row(row):
    return all(v is None or (isinstance(v, str) and v == "") for v in row)


def convert_xlsx_to_csv(xlsx_path: Path, csv_path: Path, sheet_name: str = None) -> int:
    """xlsx 파일을 CSV로 변환. 처리한 데이터 행 수 반환."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)

    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"시트 '{sheet_name}' 을 찾을 수 없습니다. 사용 가능: {wb.sheetnames}")
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    csv_path.parent.mkdir(parents=True, exist_ok=True)

    row_count = 0
    with open(csv_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)
        for row in ws.iter_rows(values_only=True):
            if is_empty_row(row):
                continue
            cleaned = [clean_cell(v) for v in row]
            writer.writerow(cleaned)
            row_count += 1

    wb.close()
    return row_count


def find_default_xlsx_files(base: Path):
    """기본 위치의 xlsx 파일 전체 탐색."""
    return sorted(base.rglob("*.xlsx"))


def main():
    parser = argparse.ArgumentParser(
        description="xlsx를 UTF-8 CSV로 변환합니다.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog=__doc__,
    )
    parser.add_argument("input", nargs="?", help="변환할 xlsx 파일 경로 (생략 시 기본 위치 전체)")
    parser.add_argument("-o", "--output", help="출력 CSV 경로 (생략 시 xlsx와 동일 위치)")
    parser.add_argument("-s", "--sheet", help="특정 시트 이름 (생략 시 첫 시트)")
    args = parser.parse_args()

    # 기본 검색 루트: 이 스크립트 기준 상위 폴더 (mods/Trans To Vostok/)
    script_dir = Path(__file__).resolve().parent
    default_base = script_dir.parent  # mods/Trans To Vostok/

    if args.input:
        xlsx_files = [Path(args.input).resolve()]
    else:
        xlsx_files = find_default_xlsx_files(default_base)
        if not xlsx_files:
            print(f"xlsx 파일을 찾을 수 없습니다: {default_base}")
            return 1

    total_rows = 0
    for xlsx in xlsx_files:
        if not xlsx.exists():
            print(f"[SKIP] 파일 없음: {xlsx}")
            continue

        if args.output and len(xlsx_files) == 1:
            csv_path = Path(args.output).resolve()
        else:
            csv_path = xlsx.with_suffix(".csv")

        try:
            rows = convert_xlsx_to_csv(xlsx, csv_path, args.sheet)
            total_rows += rows
            print(f"[OK] {xlsx.relative_to(default_base) if default_base in xlsx.parents else xlsx.name}")
            print(f"     -> {csv_path} ({rows}행)")
        except Exception as e:
            print(f"[FAIL] {xlsx}: {e}")
            return 1

    print(f"\n완료: 총 {total_rows}행 변환")
    return 0


if __name__ == "__main__":
    sys.exit(main())
