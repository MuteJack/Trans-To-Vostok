"""Phase 3 step — extract texture rework credits from Texture.xlsx -> credits.json.

Reads the `Reworked by` and `Contributors` columns from each sheet of
`Translations/<locale>/Texture.xlsx` (MetaData sheet skipped), dedupes,
and writes the unified list to `Trans To Vostok/<locale>/credits.json`'s
`Texture_reworker` field.

Other credits.json fields (Translator, translation_updated) are
preserved — this script only owns Texture_reworker. Crowdin/translator
fields are owned by tools/crowdin/get_member_list.py.

Texture.xlsx is the authoring source for image credits because texture
work happens locally with image-editing tools, off-Crowdin.

Input convention (Phase 3):
  <locale>   single locale
  all        every locale with enabled=true (English/Template excluded)
  Template   rejected — Template is never built.

Usage:
  python tools/build/get_texture_credits.py Korean
  python tools/build/get_texture_credits.py all
"""
import argparse
import json
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install -r tools/requirements.txt",
          file=sys.stderr)
    sys.exit(1)


SCRIPT_DIR = Path(__file__).resolve().parent
TOOLS_DIR = SCRIPT_DIR.parent
REPO = TOOLS_DIR.parent

sys.path.insert(0, str(TOOLS_DIR))
from helper.helper_locale_config import load_locales  # noqa: E402

MOD_ROOT = REPO / "Trans To Vostok"
DRY_RUN_ROOT = REPO / ".tmp" / "temp_build" / "Trans To Vostok"
TRANSLATIONS_ROOT = REPO / "Translations"
CREDITS_FILE = "credits.json"
TEXTURE_XLSX = "Texture.xlsx"
TEXTURE_REWORKED_COLUMN = "Reworked by"
TEXTURE_CONTRIBUTORS_COLUMN = "Contributors"
TEMPLATE_LOCALE = "Template"
SOURCE_LOCALE = "English"


def _say(msg: str = "") -> None:
    print(msg, flush=True)


def _split_names(value) -> list[str]:
    """Split a cell value into trimmed names. Excel allows in-cell newlines (Alt+Enter)."""
    if value is None:
        return []
    if not isinstance(value, str):
        s = str(value).strip()
        return [s] if s else []
    return [n.strip() for n in value.replace("\r\n", "\n").split("\n") if n.strip()]


def collect_texture_credits(xlsx_path: Path) -> list[str]:
    """Ordered, deduped list of names from `Reworked by` + `Contributors` across all sheets."""
    if not xlsx_path.exists():
        return []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    except Exception as e:
        print(f"  [WARN] Cannot read {xlsx_path}: {e}", file=sys.stderr)
        return []

    seen: set[str] = set()
    out: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue
            header = [str(c).strip() if c is not None else "" for c in rows[0]]
            rb_idx = header.index(TEXTURE_REWORKED_COLUMN) if TEXTURE_REWORKED_COLUMN in header else None
            co_idx = header.index(TEXTURE_CONTRIBUTORS_COLUMN) if TEXTURE_CONTRIBUTORS_COLUMN in header else None
            for row in rows[1:]:
                if row is None:
                    continue
                for idx in (rb_idx, co_idx):
                    if idx is None or idx >= len(row):
                        continue
                    for name in _split_names(row[idx]):
                        if name not in seen:
                            seen.add(name)
                            out.append(name)
    finally:
        wb.close()
    return out


def update_credits_for_locale(locale: str, dry_run: bool = False) -> tuple[int, list[str]]:
    """Read Texture.xlsx, write Texture_reworker field. Preserves other fields.

    Returns (exit_code, names).
    """
    xlsx = TRANSLATIONS_ROOT / locale / TEXTURE_XLSX
    if not xlsx.exists():
        _say(f"  [SKIP] Texture.xlsx not found: {xlsx.relative_to(REPO)}")
        return 0, []

    names = collect_texture_credits(xlsx)

    # input base from deploy; output to dry root when --dry-run
    base_path = MOD_ROOT / locale / CREDITS_FILE
    out_root = DRY_RUN_ROOT if dry_run else MOD_ROOT
    credits_path = out_root / locale / CREDITS_FILE
    existing: dict = {}
    if base_path.exists():
        try:
            existing = json.loads(base_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError:
            _say(f"  [WARN] {base_path.relative_to(REPO)} is not valid JSON; recreating from scratch.")
            existing = {}

    existing["Texture_reworker"] = names

    credits_path.parent.mkdir(parents=True, exist_ok=True)
    credits_path.write_text(
        json.dumps(existing, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    if names:
        _say(f"  Texture_reworker: {len(names)} ({', '.join(names)})")
    else:
        _say("  Texture_reworker: 0 (none)")
    _say(f"  -> {credits_path.relative_to(REPO)}")
    return 0, names


def _discover_enabled_locales() -> list[str]:
    out: list[str] = []
    for loc in load_locales():
        name = loc.get("dir")
        if loc.get("enabled") and name and name not in (SOURCE_LOCALE, TEMPLATE_LOCALE):
            out.append(name)
    return out


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("locale", help="Locale name or 'all'")
    parser.add_argument("--dry-run", action="store_true",
                        help="Output to .tmp/temp_build/Trans To Vostok/<locale>/ instead of deploy path")
    args = parser.parse_args(argv[1:])

    if args.locale == TEMPLATE_LOCALE:
        _say(f"[ERROR] {TEMPLATE_LOCALE} is not a valid argument for this build step.")
        _say("        Template is never built — only validated.")
        return 1

    if args.locale == "all":
        locales = _discover_enabled_locales()
        if not locales:
            _say("[ERROR] no enabled locales found in src/locale.json")
            return 1
    else:
        if not (TRANSLATIONS_ROOT / args.locale).exists():
            _say(f"[ERROR] locale directory not found: Translations/{args.locale}")
            return 1
        locales = [args.locale]

    _say(f"=== get_texture_credits ({len(locales)} locale(s)) ===")
    _say(f"  Locales: {', '.join(locales)}")
    _say()

    overall = 0
    for loc in locales:
        _say(f"--- {loc} ---")
        rc, _ = update_credits_for_locale(loc, dry_run=args.dry_run)
        if rc != 0:
            overall = rc
        _say()

    return overall


if __name__ == "__main__":
    sys.exit(main(sys.argv))
