"""Phase 3 step — generate Texture_Attribution.md from Texture.xlsx.

This file lists ONLY third-party data source attributions for image
assets (the `Attribution` column of Texture.xlsx). Person-level credit
(Reworked by / Contributors) lives in Translation_Credit.md, generated
by build_translation_credit.py.

Required columns per sheet:
    - File Name
    - Attribution

If Texture.xlsx is absent for a locale, the step is a no-op (locales
without texture translation are normal).

Input convention (Phase 3):
  <locale>   single locale
  all        every locale with enabled=true (English/Template excluded)
  Template   rejected — Template is never built.

Usage:
  python tools/build/build_attributions.py Korean
  python tools/build/build_attributions.py all
"""
import argparse
import re
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install openpyxl", file=sys.stderr)
    sys.exit(1)


SCRIPT_DIR = Path(__file__).resolve().parent
TOOLS_DIR = SCRIPT_DIR.parent
REPO = TOOLS_DIR.parent

sys.path.insert(0, str(TOOLS_DIR))
from helper.helper_locale_config import load_locales  # noqa: E402
from helper.helper_log import setup_logpath  # noqa: E402

MOD_ROOT = REPO / "Trans To Vostok"
DRY_RUN_ROOT = REPO / ".tmp" / "temp_build" / "Trans To Vostok"
TRANSLATIONS_ROOT = REPO / "Translations"
TEXTURE_XLSX = "Texture.xlsx"
OUTPUT_NAME = "Texture_Attribution.md"
TEMPLATE_LOCALE = "Template"
SOURCE_LOCALE = "English"

URL_RE = re.compile(r'https?://[^\s)\]]+')
REQUIRED_COLUMNS = ["File Name", "Attribution"]


def _say(msg: str = "") -> None:
    print(msg, flush=True)


def collect_rows(xlsx_path: Path) -> list[dict]:
    """Collect (sheet, file_name, attribution) rows from all sheets."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    rows: list[dict] = []
    try:
        for ws in wb.worksheets:
            header = [c.value for c in ws[1]]
            try:
                idx = {col: header.index(col) for col in REQUIRED_COLUMNS}
            except ValueError as e:
                _say(f"  [WARN] '{ws.title}' sheet missing required column ({e}), skipping")
                continue

            for row_no, row in enumerate(ws.iter_rows(values_only=True), 1):
                if row_no == 1:
                    continue

                def cell(name: str) -> str:
                    i = idx[name]
                    return str(row[i]).strip() if i < len(row) and row[i] is not None else ""

                file_name = cell("File Name")
                if not file_name:
                    continue

                rows.append({
                    "sheet": ws.title,
                    "file_name": file_name,
                    "attribution": cell("Attribution"),
                })
    finally:
        wb.close()
    return rows


def linkify(text: str) -> str:
    return URL_RE.sub(lambda m: f"<{m.group(0)}>", text)


def render_markdown(rows: list[dict], locale: str) -> str:
    with_attr = [r for r in rows if r["attribution"]]
    without_attr = [r for r in rows if not r["attribution"]]

    lines: list[str] = []
    lines.append(f"# Attribution - Trans To Vostok ({locale})")
    lines.append("")
    lines.append(
        "This document lists attributions for translated image assets in this mod."
    )
    lines.append("")
    lines.append(
        f"_Auto-generated from `{locale}/Texture.xlsx` by "
        "`tools/build/build_attributions.py`. Do not edit manually - "
        "update the xlsx instead._"
    )
    lines.append("")

    if with_attr:
        lines.append("## Files with third-party attribution")
        lines.append("")
        for r in with_attr:
            lines.append(f"### `{r['file_name']}` _({r['sheet']})_")
            lines.append("")
            lines.append("- **Sources**:")
            for src_line in r["attribution"].splitlines():
                src_line = src_line.strip()
                if not src_line:
                    continue
                lines.append(f"  - {linkify(src_line)}")
            lines.append("")

    if without_attr:
        lines.append("## Files without third-party attribution")
        lines.append("")
        lines.append(
            "Original work without third-party sources. "
            "For per-language contributor credit, see `Translation_Credit.md`."
        )
        lines.append("")
        by_sheet: dict[str, list[dict]] = {}
        for r in without_attr:
            by_sheet.setdefault(r["sheet"], []).append(r)
        for sheet_name in by_sheet:
            lines.append(f"### {sheet_name}")
            lines.append("")
            for r in by_sheet[sheet_name]:
                lines.append(f"- `{r['file_name']}`")
            lines.append("")

    if not rows:
        lines.append("_No image assets registered._")
        lines.append("")

    return "\n".join(lines)


def build_for_locale(locale: str, dry_run: bool = False) -> int:
    xlsx_path = TRANSLATIONS_ROOT / locale / TEXTURE_XLSX
    if not xlsx_path.exists():
        _say(f"  [SKIP] Texture.xlsx not found: {xlsx_path.relative_to(REPO)}")
        return 0

    rows = collect_rows(xlsx_path)
    sheets = sorted({r["sheet"] for r in rows})
    _say(f"  Loaded {len(rows)} entries from sheets: {', '.join(sheets) or '-'}")

    md = render_markdown(rows, locale)
    out_root = DRY_RUN_ROOT if dry_run else MOD_ROOT
    out_path = out_root / locale / OUTPUT_NAME
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(md, encoding="utf-8")

    with_attr = sum(1 for r in rows if r["attribution"])
    without_attr = len(rows) - with_attr
    _say(f"  with attribution:    {with_attr}")
    _say(f"  without attribution: {without_attr}")
    _say(f"  -> {out_path.relative_to(REPO)}")
    return 0


def _discover_enabled_locales() -> list[str]:
    out: list[str] = []
    for loc in load_locales():
        name = loc.get("dir")
        if loc.get("enabled") and name and name not in (SOURCE_LOCALE, TEMPLATE_LOCALE):
            out.append(name)
    return out


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("locale", help="Locale name or 'all'")
    parser.add_argument("--dry-run", action="store_true",
                        help="Output to .tmp/temp_build/Trans To Vostok/<locale>/ instead of deploy path")
    parser.add_argument("--logpath", default=None,
                        help="Append stdout/stderr to this log file "
                             "(used by orchestrator)")
    args = parser.parse_args(argv[1:])
    setup_logpath(args.logpath)

    if args.locale == TEMPLATE_LOCALE:
        _say(f"[ERROR] {TEMPLATE_LOCALE} is not a valid argument for this build step.")
        _say("        Template is never built - only validated.")
        return 1

    if args.locale == "all":
        locales = _discover_enabled_locales()
        if not locales:
            _say("[ERROR] no enabled locales found in src/locale.json")
            return 1
    else:
        if not (TRANSLATIONS_ROOT / args.locale).exists():
            _say(f"[ERROR] locale directory not found: Translations/{args.locale}")
            return 1
        locales = [args.locale]

    _say(f"=== build_attributions ({len(locales)} locale(s)) ===")
    _say(f"  Locales: {', '.join(locales)}")
    _say()

    overall = 0
    for loc in locales:
        _say(f"--- {loc} ---")
        rc = build_for_locale(loc, dry_run=args.dry_run)
        if rc != 0:
            overall = rc
        _say()

    return overall


if __name__ == "__main__":
    sys.exit(main(sys.argv))
