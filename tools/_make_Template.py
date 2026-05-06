"""Create a blank translation template from an existing locale's Translation.xlsx.

Copies <source>/Translation.xlsx into the mod's .tmp/ folder, then on the
copy:
  - clears the `translation` column for every row
  - sets the boolean-style status columns (Transliteration / Transcreation /
    Machine translated / Confused / untranslatable) to 0
  - keeps the source `text` and all metadata columns (method, filename,
    location, parent, name, type, property, unique_id, …) untouched

Used as the starting point when bootstrapping a new locale.

Usage:
    python tools/_make_Template.py
    python tools/_make_Template.py --source Korean
    python tools/_make_Template.py --output my/path.xlsx
"""

import argparse
import shutil
import sys
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent  # mod root (Trans To Vostok mod)
DEFAULT_SOURCE_LOCALE = "Korean"
DEFAULT_OUTPUT = ROOT / ".tmp" / "Translation_template.xlsx"

CLEAR_COLUMN = "translation"
ZERO_COLUMNS = [
    "Transliteration",
    "Transcreation",
    "Machine translated",
    "Confused",
    "untranslatable",
]


def make_template(source_xlsx: Path, output_xlsx: Path) -> None:
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(source_xlsx, output_xlsx)
    print(f"Copied: {source_xlsx} -> {output_xlsx}")

    wb = openpyxl.load_workbook(output_xlsx)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ws.max_row < 2:
            print(f"  [skip] {sheet_name}: no data rows")
            continue

        header = {}
        for c in ws[1]:
            if c.value is not None:
                header[str(c.value)] = c.column

        clear_idx = header.get(CLEAR_COLUMN)
        zero_indices = [header[col] for col in ZERO_COLUMNS if col in header]

        if clear_idx is None and not zero_indices:
            print(f"  [skip] {sheet_name}: no relevant columns")
            continue

        for row_idx in range(2, ws.max_row + 1):
            if clear_idx is not None:
                ws.cell(row_idx, clear_idx).value = None
            for zi in zero_indices:
                ws.cell(row_idx, zi).value = 0

        rows = ws.max_row - 1
        actions = []
        if clear_idx is not None:
            actions.append(f"cleared `{CLEAR_COLUMN}`")
        if zero_indices:
            actions.append(f"zeroed {len(zero_indices)} status columns")
        print(f"  {sheet_name}: {' + '.join(actions)} ({rows} rows)")

    wb.save(output_xlsx)
    print(f"\nTemplate written: {output_xlsx}")


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__,
                                formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument(
        "--source",
        default=DEFAULT_SOURCE_LOCALE,
        help=f"Source locale folder under 'Trans To Vostok/' (default: {DEFAULT_SOURCE_LOCALE})",
    )
    p.add_argument(
        "--output",
        type=Path,
        default=None,
        help=f"Output xlsx path (default: {DEFAULT_OUTPUT.relative_to(ROOT)})",
    )
    args = p.parse_args()

    source_xlsx = ROOT / "Trans To Vostok" / args.source / "Translation.xlsx"
    if not source_xlsx.exists():
        print(f"ERROR: source not found: {source_xlsx}", file=sys.stderr)
        return 1

    output = args.output if args.output else DEFAULT_OUTPUT
    make_template(source_xlsx, output)
    return 0


if __name__ == "__main__":
    sys.exit(main())
