"""
Build a draft translation glossary from Translation.xlsx.

Strategy (no external API needed):
1. Treat short entries (<=4 words) as direct en->ko candidate pairs.
2. Aggregate identical entries; count frequency.
3. Extract single-word frequency from longer entries (descriptions).
4. Group by length / category for human review.
5. Write TSV. Humans should curate the result.

Output:
    Korean/glossary_draft.tsv

Usage:
    python tools/build_glossary.py
    python tools/build_glossary.py --locale Korean
    python tools/build_glossary.py --min-freq 2
"""
import argparse
import csv
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install openpyxl", file=sys.stderr)
    sys.exit(1)

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass


SKIP_SHEETS = {"MetaData"}
SHORT_MAX_WORDS = 4
SHORT_MAX_CHARS = 32

BOOL_TRUE = {"1", "true"}

# Words that are usually not glossary-worthy (English stopwords)
STOPWORDS = {
    "the", "a", "an", "of", "to", "in", "on", "at", "by", "for", "with",
    "and", "or", "but", "is", "are", "was", "were", "be", "been", "being",
    "have", "has", "had", "do", "does", "did", "will", "would", "should",
    "can", "could", "may", "might", "must", "shall",
    "i", "you", "he", "she", "it", "we", "they", "me", "him", "her", "us",
    "them", "my", "your", "his", "its", "our", "their",
    "this", "that", "these", "those", "what", "which", "who", "whom",
    "if", "then", "else", "as", "than", "so", "not", "no", "yes",
    "from", "into", "out", "up", "down", "over", "under", "off", "out",
    "very", "more", "most", "less", "much", "many", "some", "any", "all",
    "one", "two", "three", "first", "second", "third",
    "now", "here", "there", "when", "where", "why", "how",
    "yourself", "another", "etc", "only", "just", "also", "too",
}

# Token regex: alphabetic words 2+ chars, optionally hyphenated
WORD_RE = re.compile(r"[A-Za-z][A-Za-z\-]{1,}")


def collect_pairs(xlsx_path: Path) -> list[tuple[str, str, str]]:
    """Return list of (text, translation, sheet) for rows with both filled."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    pairs: list[tuple[str, str, str]] = []
    for ws in wb.worksheets:
        if ws.title in SKIP_SHEETS:
            continue
        header = [c.value for c in ws[1]]
        try:
            i_text = header.index("text")
            i_trans = header.index("translation")
        except ValueError:
            continue
        i_method = header.index("method") if "method" in header else None
        i_untranslatable = header.index("untranslatable") if "untranslatable" in header else None

        for row_no, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_no == 1:
                continue
            if i_text >= len(row) or i_trans >= len(row):
                continue
            text = row[i_text]
            trans = row[i_trans]
            if not text or not trans:
                continue
            method = (str(row[i_method]).strip().lower()
                      if i_method is not None and i_method < len(row) and row[i_method] is not None
                      else "")
            if method == "ignore":
                continue
            untrans = (str(row[i_untranslatable]).strip().lower()
                       if i_untranslatable is not None and i_untranslatable < len(row) and row[i_untranslatable] is not None
                       else "")
            if untrans in BOOL_TRUE:
                continue
            pairs.append((str(text).strip(), str(trans).strip(), ws.title))
    return pairs


def is_short(text: str) -> bool:
    """Short entries are good direct glossary candidates."""
    if len(text) > SHORT_MAX_CHARS:
        return False
    word_count = len(WORD_RE.findall(text))
    return 0 < word_count <= SHORT_MAX_WORDS


def normalize_short(text: str) -> str:
    """Comparison key for short entries (whitespace-collapsed, lowercased)."""
    return " ".join(text.split()).lower()


def build_short_glossary(
    pairs: list[tuple[str, str, str]], min_freq: int = 1
) -> list[dict]:
    """
    Aggregate short (text, translation) pairs.

    Returns a list of dicts sorted by frequency desc.
    Conflicting Korean for same English text are listed in 'variants'.
    """
    grouped: dict[str, dict] = {}
    for text, trans, sheet in pairs:
        if not is_short(text):
            continue
        key = normalize_short(text)
        if key not in grouped:
            grouped[key] = {
                "en_canonical": text,
                "translations": Counter(),
                "sheets": Counter(),
                "count": 0,
            }
        grouped[key]["translations"][trans] += 1
        grouped[key]["sheets"][sheet] += 1
        grouped[key]["count"] += 1

    out: list[dict] = []
    for key, data in grouped.items():
        if data["count"] < min_freq:
            continue
        # most common translation
        ko_top, ko_top_count = data["translations"].most_common(1)[0]
        variants = [
            f"{t} ({c})" for t, c in data["translations"].most_common()
            if t != ko_top
        ]
        out.append({
            "en": data["en_canonical"],
            "ko": ko_top,
            "freq": data["count"],
            "ko_consistency": f"{ko_top_count}/{data['count']}",
            "variants": "; ".join(variants),
            "sheets": ", ".join(s for s, _ in data["sheets"].most_common()),
        })
    out.sort(key=lambda d: (-d["freq"], d["en"].lower()))
    return out


def build_word_frequency(
    pairs: list[tuple[str, str, str]], min_freq: int = 3
) -> list[dict]:
    """
    Word-level frequency on long source texts.
    Useful for surfacing recurring vocabulary in descriptions.

    Capitalised tokens are flagged as likely proper nouns.
    """
    word_counter: Counter = Counter()
    proper_counter: Counter = Counter()
    for text, _trans, _sheet in pairs:
        if is_short(text):
            continue  # skip short entries (already covered)
        for token in WORD_RE.findall(text):
            lower = token.lower()
            if lower in STOPWORDS:
                continue
            word_counter[lower] += 1
            if token[0].isupper() and not text.startswith(token):
                # capitalised mid-sentence -> likely proper noun
                proper_counter[token] += 1

    out: list[dict] = []
    for w, c in word_counter.most_common():
        if c < min_freq:
            break
        is_proper = w in {p.lower() for p in proper_counter}
        out.append({
            "word": w,
            "freq": c,
            "type": "proper" if is_proper else "common",
        })
    return out


def write_tsv(out_path: Path, short_glossary: list[dict], word_freq: list[dict]) -> None:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = out_path.with_suffix(out_path.suffix + ".tmp")
    try:
        with open(tmp_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)

            writer.writerow(["# === SHORT ENTRIES (direct mapping candidates) ==="])
            writer.writerow(["en", "ko", "freq", "ko_consistency", "variants", "sheets"])
            for row in short_glossary:
                writer.writerow([
                    row["en"], row["ko"], row["freq"],
                    row["ko_consistency"], row["variants"], row["sheets"],
                ])

            writer.writerow([])
            writer.writerow(["# === WORD FREQUENCY (long-text vocabulary) ==="])
            writer.writerow(["word", "freq", "type"])
            for row in word_freq:
                writer.writerow([row["word"], row["freq"], row["type"]])

        tmp_path.replace(out_path)
    except Exception:
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                pass
        raise


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Build a draft translation glossary from Translation.xlsx"
    )
    parser.add_argument("--locale", default="Korean")
    parser.add_argument("--min-freq", type=int, default=1,
                        help="Minimum frequency for short entries (default 1)")
    parser.add_argument("--word-min-freq", type=int, default=3,
                        help="Minimum frequency for word-level analysis (default 3)")
    parser.add_argument("--output", default=None,
                        help="Output path (default: <pkg_root>/<locale>/glossary_draft.tsv)")
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    mod_root = script_dir.parent
    pkg_root = mod_root / "Trans To Vostok"
    xlsx_path = pkg_root / args.locale / "Translation.xlsx"

    if not xlsx_path.exists():
        print(f"ERROR: Translation.xlsx not found: {xlsx_path}", file=sys.stderr)
        return 1

    print(f"Input:  {xlsx_path}")
    pairs = collect_pairs(xlsx_path)
    print(f"Loaded: {len(pairs)} (text, translation) pairs")

    short_glossary = build_short_glossary(pairs, min_freq=args.min_freq)
    word_freq = build_word_frequency(pairs, min_freq=args.word_min_freq)

    out_path = (
        Path(args.output).resolve()
        if args.output
        else (pkg_root / args.locale / "glossary_draft.tsv")
    )
    write_tsv(out_path, short_glossary, word_freq)
    print(f"Output: {out_path}")
    print()
    print(f"  Short-entry candidates : {len(short_glossary)}")
    print(f"  Word-frequency entries : {len(word_freq)}")

    # ratio of consistent translations
    consistent = sum(1 for r in short_glossary
                     if r["ko_consistency"].split("/")[0]
                     == r["ko_consistency"].split("/")[1])
    if short_glossary:
        ratio = consistent / len(short_glossary) * 100
        print(f"  Of short entries, {consistent}/{len(short_glossary)} ({ratio:.1f}%) "
              "have a single consistent Korean translation.")
        inconsistent = len(short_glossary) - consistent
        if inconsistent:
            print(f"  {inconsistent} entries have multiple translations — see 'variants' column.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
