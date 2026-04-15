"""
[임시 도구] Translation.xlsx 의 번역 시트들을 각각 TSV로 덤프.

디버깅/검토용. 엑셀 스키마 변경 중 내용 확인 및 diff 비교에 쓴다.
이 도구는 xlsx 를 수정하지 않고 읽기 전용으로 연다.

MetaData 시트는 번역 데이터가 아니라 설정 시트이므로 덤프 대상에서 제외한다.
나머지 시트마다 시트명을 파일명에 붙여 개별 TSV 파일을 만든다.

    Main       → .tmp/_Main.tsv
    Interface  → .tmp/_Interface.tsv
    (MetaData) → 스킵

사용법:
    python _xlsx_to_tsv.py                      # 기본: Korean/Translation.xlsx
    python _xlsx_to_tsv.py <xlsx_path>          # 입력 지정
    python _xlsx_to_tsv.py <xlsx_path> <out_dir>  # 입력 + 출력 디렉토리 지정

기본 출력: .tmp/_{Sheet}.tsv
(Korean/translation.tsv 런타임 빌드 산출물과의 충돌 회피)
"""
import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl이 필요합니다. pip install openpyxl", file=sys.stderr)
    sys.exit(1)

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass


# 덤프 제외 시트 (번역 데이터가 아님)
SKIP_SHEETS = {"MetaData"}

# 엑셀 아티팩트 제거용
_X000D_RE = re.compile(r"_x000[dD]_")


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    s = str(value)
    s = _X000D_RE.sub("", s)
    return s


def _normalize_header(value) -> str:
    s = _normalize_cell(value)
    s = s.replace("\r", " ").replace("\n", " ")
    return " ".join(s.split())


def dump_sheet(ws, out_path: Path) -> tuple[int, int]:
    """한 시트를 TSV 파일로 원자적으로 쓴다. 반환: (컬럼 수, 데이터 행 수)"""
    rows_iter = ws.iter_rows(values_only=True)
    header_raw = next(rows_iter, None)
    if header_raw is None:
        return (0, 0)
    header = [_normalize_header(c) for c in header_raw]

    data_rows: list[list[str]] = []
    for row_values in rows_iter:
        if row_values is None or all(v is None for v in row_values):
            continue
        data_rows.append([_normalize_cell(v) for v in row_values])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = out_path.with_suffix(out_path.suffix + ".tmp")
    with open(tmp_path, "w", encoding="utf-8", newline="") as f:
        writer = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)
        writer.writerow(header)
        for row in data_rows:
            writer.writerow(row)
    tmp_path.replace(out_path)

    return (len(header), len(data_rows))


def dump_xlsx(xlsx_path: Path, out_dir: Path) -> int:
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        sheets = [name for name in wb.sheetnames if name not in SKIP_SHEETS]
        if not sheets:
            print(f"[ERROR] 덤프할 시트가 없습니다 (SKIP: {SKIP_SHEETS})", file=sys.stderr)
            return 1

        print(f"입력: {xlsx_path}")
        print(f"출력 디렉토리: {out_dir}")
        print(f"대상 시트: {sheets}  (스킵: {sorted(SKIP_SHEETS)})")
        print()

        for name in sheets:
            ws = wb[name]
            out_path = out_dir / f"_{name}.tsv"
            cols, rows = dump_sheet(ws, out_path)
            if cols == 0:
                print(f"  [WARN] {name}: 빈 시트 (스킵)")
                continue
            print(f"  [OK] {name} → {out_path.name}  ({cols}컬럼, {rows}행)")
        return 0
    finally:
        wb.close()


def main() -> int:
    script_dir = Path(__file__).resolve().parent
    mod_root = script_dir.parent

    default_xlsx = mod_root / "Korean" / "Translation.xlsx"
    default_out = mod_root / ".tmp"

    xlsx_path = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else default_xlsx.resolve()
    out_dir = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else default_out.resolve()

    if not xlsx_path.exists():
        print(f"[ERROR] xlsx 파일이 없습니다: {xlsx_path}", file=sys.stderr)
        return 1

    return dump_xlsx(xlsx_path, out_dir)


if __name__ == "__main__":
    sys.exit(main())
