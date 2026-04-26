"""
Translation.xlsx validation tool.

Schema (18 columns):
    A. Meta (not used by tool): WHERE, SUB, KIND
    B. Status flags (validation only): Transliteration, Machine translated, Confused, ignore
    C. Matching: method, filename, filetype, location, parent, name, type, unique_id
    D. Content: text, translation
    E. Notes (not used by tool): DESCRIPTION

method values:
    static   — (location, parent, name, type, text) exact 1:1 match (TSV-validated)
    literal  — exact text match (scoped if location set, otherwise global)
    pattern  — regex match (scoped if location set, otherwise global)
    ""       — empty value defaults to literal (manual-entry convenience)

Checks:
1. [ERROR] required columns missing
2. [ERROR] TSV match (method=static only):
   - For filetype in {tscn, scn}, match against TSV by unique_id
   - text must match exactly including leading/trailing whitespace/newlines (strict)
3. [WARNING] text ↔ translation leading/trailing whitespace/newline mismatch
4. [ERROR] flag value validation: Transliteration, Machine translated, Confused, ignore
   - allowed: 0/1/true/false/""
5. [ERROR] method / field combination validation:
   - static: location/parent/name/type/unique_id required, filetype∈{tscn,scn}
   - literal + location: parent/name/type required (scoped)
   - literal + no location: free context (global)
   - pattern + location: parent/name/type required (scoped)
   - pattern + no location: free context (global)
   - "": treated as literal
6. [WARNING] empty method + unique_id filled → recommend explicitly setting static
7. [ERROR] duplicate matching keys:
   - static + scoped literal: shared (location, parent, name, type, text) space
   - global literal: text space
   - scoped pattern: (location, parent, name, type, text) space
   - global pattern: text space

Usage:
    python validate_translation.py <locale>

Example:
    python validate_translation.py Korean

Log:
    <xlsx_parent_folder>/.log/validate_translation_YYYYMMDD_HHMMSS.log

Exit codes:
    0 = no errors (warnings ignored)
    1 = errors found
"""
import csv
import re
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. Please run the following command.\n"
          " >> pip install openpyxl", file=sys.stderr)
    sys.exit(1)

# Windows console Korean output support
if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass


# allowed flag values (empty value also allowed — treated as unspecified)
VALID_FLAGS = {"", "0", "1", "true", "false"}

# flag columns (ignore is replaced by method=ignore)
FLAG_COLUMNS = ["Transliteration", "Machine translated", "Confused", "untranslatable"]

# valid method values (empty string defaults to literal; ignore is treated as exclusion)
VALID_METHODS = {"", "static", "literal", "pattern", "substr", "ignore"}

# scene file extensions (source types)
SCENE_FILETYPES = {"tscn", "scn"}

# exclude MetaData sheet (not translation data)
SKIP_SHEETS = {"MetaData"}

# for stripping the Excel _x000D_ artifact
_X000D_RE = re.compile(r"_x000[dD]_")


# ==========================================
# utilities
# ==========================================

def _normalize_cell(value) -> str:
    """Normalize an Excel cell value to a string. Strips _x000d_ / _x000D_ artifacts."""
    if value is None:
        return ""
    s = str(value)
    s = _X000D_RE.sub("", s)
    return s


def _leading_ws(s: str) -> str:
    return s[: len(s) - len(s.lstrip())]


def _trailing_ws(s: str) -> str:
    return s[len(s.rstrip()):]


def _preview(text: str, limit: int = 60) -> str:
    r = repr(text)
    if len(r) > limit:
        return r[: limit - 3] + "...'"
    return r


def _effective_method(row: dict) -> str:
    """Treat empty method as literal by default. ignore is returned as-is."""
    m = row.get("method", "").strip()
    if m == "ignore":
        return "ignore"
    return m if m else "literal"


# ==========================================
# TSV index (built from extracted .tscn.tsv / .tres.tsv by unique_id)
# ==========================================

def load_tsv_index(tsv_dir: Path) -> dict:
    """
    Build a unique_id → [list of records] index from extracted TSV files.

    Target files:
        *.tscn.tsv  (output of parse_tscn_text.py, has unique_id)
        *.tres.tsv  (output of parse_tres_text.py, no unique_id → naturally skipped)
    """
    index = {}
    tsv_files = sorted(tsv_dir.rglob("*.tscn.tsv")) + sorted(tsv_dir.rglob("*.tres.tsv"))
    for tsv_file in tsv_files:
        try:
            with open(tsv_file, "r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f, delimiter="\t")
                for row in reader:
                    uid = (row.get("unique_id") or "").strip()
                    if not uid:
                        continue
                    record = {
                        "filename": (row.get("filename") or "").strip(),
                        "filetype": (row.get("filetype") or "").strip(),
                        "location": (row.get("location") or "").strip(),
                        "parent": (row.get("parent") or "").strip(),
                        "name": (row.get("name") or "").strip(),
                        "type": (row.get("type") or "").strip(),
                        "text": row.get("text") or "",
                        "_tsv_file": tsv_file.name,
                    }
                    index.setdefault(uid, []).append(record)
        except Exception as e:
            print(f"[WARN] Failed to read TSV: {tsv_file} ({e})")
    return index


def load_tres_text_set(tsv_dir: Path) -> set:
    """
    Collect all text values from extracted *.tres.tsv files into a set.
    Used to verify that tres xlsx row texts exist in the actual .tres source.
    """
    texts: set = set()
    for tsv_file in sorted(tsv_dir.rglob("*.tres.tsv")):
        try:
            with open(tsv_file, "r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f, delimiter="\t")
                for row in reader:
                    text = row.get("text") or ""
                    if text:
                        texts.add(text)
        except Exception as e:
            print(f"[WARN] Failed to read tres TSV: {tsv_file} ({e})")
    return texts


def load_gd_text_set(tsv_dir: Path) -> set:
    """
    Collect all text values from extracted *.gd.tsv files into a set.
    Used to verify that filetype=gd xlsx row texts exist in the actual .gd source.
    """
    texts: set = set()
    for tsv_file in sorted(tsv_dir.rglob("*.gd.tsv")):
        try:
            with open(tsv_file, "r", encoding="utf-8", newline="") as f:
                reader = csv.DictReader(f, delimiter="\t")
                for row in reader:
                    text = row.get("text") or ""
                    if text:
                        texts.add(text)
        except Exception as e:
            print(f"[WARN] Failed to read gd TSV: {tsv_file} ({e})")
    return texts


# ==========================================
# MetaData loading
# ==========================================

def load_metadata(xlsx_path: Path) -> dict:
    """
    Return a Field → Value dict from the MetaData sheet.
    Returns an empty dict if MetaData sheet is absent.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        if "MetaData" not in wb.sheetnames:
            return {}
        ws = wb["MetaData"]
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, None)
        if header is None:
            return {}
        meta: dict = {}
        for row in rows_iter:
            if row is None or len(row) < 3:
                continue
            field = _normalize_cell(row[1])
            value = _normalize_cell(row[2])
            if field:
                meta[field] = value
        return meta
    finally:
        wb.close()


def format_metadata_lines(meta: dict) -> list[str]:
    """Return metadata as a list of strings for log output."""
    lines = []
    for key in ("Game Version", "Game Updated Date", "Translation Updated Date",
                "Translation Updated Time", "Translation UTC", "Translator"):
        val = meta.get(key, "")
        if val:
            lines.append(f"  {key}: {val}")
    return lines


# ==========================================
# xlsx loading
# ==========================================

def load_xlsx_main(xlsx_path: Path, sheet_name: str = "Main") -> tuple[list[str], list[dict], str]:
    """
    Read the specified sheet from Translation.xlsx and return (header, list of rows, sheet name).
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"'{sheet_name}' sheet not found: {wb.sheetnames}")
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    header = []
    for c in next(rows_iter):
        h = _normalize_cell(c)
        h = h.replace("\r", " ").replace("\n", " ")
        h = " ".join(h.split())
        header.append(h)

    rows = []
    for row_values in rows_iter:
        if row_values is None or all(v is None for v in row_values):
            continue
        row_dict = {}
        for i, key in enumerate(header):
            val = row_values[i] if i < len(row_values) else None
            # if the same column name occurs multiple times, only the first is used (rest discarded)
            if key not in row_dict:
                row_dict[key] = _normalize_cell(val)
        rows.append(row_dict)

    wb.close()
    return header, rows, sheet_name


def load_all_translation_sheets(xlsx_path: Path) -> list[tuple[str, list[str], list[dict]]]:
    """
    Load all sheets except MetaData and return as a list of (sheet_name, header, rows) tuples.
    """
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        result = []
        for sheet_name in wb.sheetnames:
            if sheet_name in SKIP_SHEETS:
                continue
            ws = wb[sheet_name]
            rows_iter = ws.iter_rows(values_only=True)
            header_raw = next(rows_iter, None)
            if header_raw is None:
                continue
            header = []
            for c in header_raw:
                h = _normalize_cell(c)
                h = h.replace("\r", " ").replace("\n", " ")
                h = " ".join(h.split())
                header.append(h)

            rows: list[dict] = []
            for row_values in rows_iter:
                if row_values is None or all(v is None for v in row_values):
                    continue
                row_dict = {"_sheet": sheet_name}
                for i, key in enumerate(header):
                    val = row_values[i] if i < len(row_values) else None
                    if key not in row_dict:
                        row_dict[key] = _normalize_cell(val)
                rows.append(row_dict)
            result.append((sheet_name, header, rows))
        return result
    finally:
        wb.close()


# ==========================================
# individual checks
# ==========================================

def check_tsv_match(row: dict, tsv_index: dict) -> list[str]:
    """
    Check 2 (ERROR): TSV match for method=static rows only.
    Requires filetype in {tscn, scn}; matched against the extracted index by unique_id.
    text must match exactly including leading/trailing whitespace/newlines (strict).
    """
    errors = []
    method = _effective_method(row)
    if method != "static":
        return errors  # only static is subject to TSV validation

    filetype = row.get("filetype", "").strip()
    if filetype not in SCENE_FILETYPES:
        return errors  # cannot validate (source is not tscn/scn)

    uid = row.get("unique_id", "").strip()
    if not uid:
        errors.append("static: unique_id is empty")
        return errors

    candidates = tsv_index.get(uid)
    if not candidates:
        errors.append(f"static: unique_id={uid} not found in extracted TSV")
        return errors

    # if there are multiple candidates, prefer the one whose text matches (handles uid duplication across scenes)
    xlsx_text = row.get("text", "")
    best = None
    for cand in candidates:
        if cand["text"] == xlsx_text:
            best = cand
            break
    if best is None:
        best = candidates[0]

    mismatches = []
    # fields to compare (xlsx → TSV field names)
    field_map = [
        ("filename", "filename"),
        ("filetype", "filetype"),
        ("location", "location"),
        ("parent", "parent"),
        ("name", "name"),
        ("type", "type"),
    ]
    for xlsx_key, tsv_key in field_map:
        xlsx_val = row.get(xlsx_key, "").strip()
        tsv_val = best.get(tsv_key, "").strip()
        if xlsx_val != tsv_val:
            mismatches.append(
                f"{xlsx_key}: xlsx={_preview(xlsx_val)} / tsv={_preview(tsv_val)}"
            )

    # text uses strict comparison (whitespace/newline included)
    if xlsx_text != best["text"]:
        mismatches.append(
            f"text: xlsx={_preview(xlsx_text)} / tsv={_preview(best['text'])}"
        )

    if mismatches:
        errors.append(
            f"unique_id={uid} ({best['_tsv_file']}) field mismatch: " + "; ".join(mismatches)
        )

    return errors


def check_tres_text(row: dict, tres_texts: set) -> list[str]:
    """
    Check 2b (ERROR): verify that text in filetype=tres rows exists in extracted .tres.tsv.
    Targets rows with filetype=tres that are not method=static. Prevents typos / phantom entries.
    """
    errors = []
    filetype = row.get("filetype", "").strip()
    if filetype != "tres":
        return errors
    effective = _effective_method(row)
    if effective == "ignore":
        return errors

    text = row.get("text", "")
    if text and text not in tres_texts:
        filename = row.get("filename", "").strip()
        errors.append(
            f"filetype=tres text not found in extracted .tres.tsv: "
            f"filename={filename!r}, text={_preview(text, 40)}"
        )
    return errors


def check_gd_text(row: dict, gd_texts: set) -> list[str]:
    """
    Check 2c (WARNING): verify that text in filetype=gd rows exists in extracted .gd.tsv.
    Since GD extraction is heuristic-based, this is a WARNING rather than ERROR.
    """
    warnings = []
    filetype = row.get("filetype", "").strip()
    if filetype != "gd":
        return warnings
    effective = _effective_method(row)
    if effective == "ignore":
        return warnings

    text = row.get("text", "")
    if text and text not in gd_texts:
        filename = row.get("filename", "").strip()
        warnings.append(
            f"filetype=gd text not found in extracted .gd.tsv (changed/removed?): "
            f"filename={filename!r}, text={_preview(text, 40)}"
        )
    return warnings


def check_whitespace(row: dict) -> list[str]:
    """Check 3 (WARNING): text ↔ translation leading/trailing whitespace/newline match."""
    warnings = []
    text = row.get("text", "")
    translation = row.get("translation", "")

    if not translation:
        return warnings

    text_lead = _leading_ws(text)
    trans_lead = _leading_ws(translation)
    if text_lead != trans_lead:
        warnings.append(
            f"Leading whitespace mismatch: text={_preview(text_lead)} / translation={_preview(trans_lead)}"
        )

    text_trail = _trailing_ws(text)
    trans_trail = _trailing_ws(translation)
    if text_trail != trans_trail:
        warnings.append(
            f"Trailing whitespace mismatch: text={_preview(text_trail)} / translation={_preview(trans_trail)}"
        )

    return warnings


def check_flags(row: dict) -> list[str]:
    """Check 4 (ERROR): flag column value validation. Empty values allowed."""
    errors = []
    for col in FLAG_COLUMNS:
        val = row.get(col, "")
        normalized = val.strip().lower()
        if normalized not in VALID_FLAGS:
            errors.append(f"{col} has invalid value: {val!r} (allowed: 0/1/true/false or empty)")
    return errors


def check_method_fields(row: dict) -> list[str]:
    """
    Check 5 (ERROR): validate method value combined with related fields.
    """
    errors = []
    method = row.get("method", "").strip()
    location = row.get("location", "").strip()
    filetype = row.get("filetype", "").strip()
    parent = row.get("parent", "").strip()
    name = row.get("name", "").strip()
    type_ = row.get("type", "").strip()
    unique_id = row.get("unique_id", "").strip()

    if method not in VALID_METHODS:
        errors.append(f"Unknown method: {method!r} (allowed: static/literal/pattern/substr/ignore or empty)")
        return errors

    effective = method if method else "literal"

    if effective == "ignore":
        return errors  # skip field combination validation for ignore rows

    if effective == "static":
        if not location:
            errors.append("static: location required")
        if filetype not in SCENE_FILETYPES:
            errors.append(
                f"static: filetype must be 'tscn' or 'scn' (current: {filetype!r})"
            )
        if not parent and not name:
            # empty parent is OK for root nodes. but both parent+name being empty is suspicious.
            pass
        if not name:
            errors.append("static: name required")
        # type can legitimately be empty for instance nodes (no type= in .tscn). validated by TSV match.
        if not unique_id:
            errors.append("static: unique_id required")

    elif effective == "literal":
        if location:
            # scoped literal: context fields required (except type, to support instance nodes)
            if not name:
                errors.append("scoped literal: name required (literal + location)")
        # global literal: no constraint other than text

    elif effective == "pattern":
        if location:
            # scoped pattern: context fields required (except type)
            if not name:
                errors.append("scoped pattern: name required (pattern + location)")
        # global pattern: no constraint other than text (regex)

    elif effective == "substr":
        # substr requires only text + translation. Global substring replacement.
        pass

    return errors


def check_empty_method(row: dict) -> list[str]:
    """Check 6 (WARNING): empty method + unique_id filled → recommend explicitly setting static."""
    warnings = []
    method = row.get("method", "").strip()
    if method:
        return warnings
    uid = row.get("unique_id", "").strip()
    if uid:
        warnings.append(
            "method is empty but unique_id is filled - please specify static explicitly"
        )
    return warnings


def check_duplicates(rows: list[dict]) -> list[tuple[int, str]]:
    """
    Check 7 (ERROR): runtime matching key duplicates.
    - static and scoped literal share the same 5-tuple key space (same runtime behavior)
    - global literal: text
    - scoped pattern: 5-tuple
    - global pattern: text
    ignore=1 is excluded.
    """
    errors = []
    exact_keys: dict[tuple, list] = {}         # static + scoped literal
    literal_global: dict[str, list] = {}
    scoped_pattern_keys: dict[tuple, list] = {}
    pattern_global: dict[str, list] = {}
    substr_global: dict[str, list] = {}

    for i, row in enumerate(rows, start=2):
        effective = _effective_method(row)
        untranslatable = row.get("untranslatable", "").strip().lower() in {"1", "true"}
        if effective == "ignore" or untranslatable:
            continue

        text = row.get("text", "")
        location = row.get("location", "").strip()

        key_5 = (
            location,
            row.get("parent", "").strip(),
            row.get("name", "").strip(),
            row.get("type", "").strip(),
            text,
        )

        if effective == "static":
            exact_keys.setdefault(key_5, []).append((i, row))
        elif effective == "literal":
            if location:
                exact_keys.setdefault(key_5, []).append((i, row))
            else:
                literal_global.setdefault(text, []).append((i, row))
        elif effective == "pattern":
            if location:
                scoped_pattern_keys.setdefault(key_5, []).append((i, row))
            else:
                pattern_global.setdefault(text, []).append((i, row))
        elif effective == "substr":
            substr_global.setdefault(text, []).append((i, row))

    def _emit(label: str, store: dict, is_tuple_key: bool):
        for key, occurrences in store.items():
            if len(occurrences) <= 1:
                continue
            row_nums = ", ".join(str(n) for n, _ in occurrences)
            if is_tuple_key:
                loc, parent, name, type_, text = key
                text_preview = _preview(text, 40)
                msg = (
                    f"{label} ({len(occurrences)} occurrences): "
                    f"location={loc!r}, parent={parent!r}, name={name!r}, type={type_!r}, "
                    f"text={text_preview} (rows: {row_nums})"
                )
            else:
                text_preview = _preview(key, 40)
                msg = (
                    f"{label} ({len(occurrences)} occurrences): "
                    f"text={text_preview} (rows: {row_nums})"
                )
            for row_num, _ in occurrences:
                errors.append((row_num, msg))

    _emit("exact match duplicate (static/scoped literal)", exact_keys, True)
    _emit("global literal duplicate", literal_global, False)
    _emit("scoped pattern duplicate", scoped_pattern_keys, True)
    _emit("global pattern duplicate", pattern_global, False)
    _emit("global substr duplicate", substr_global, False)

    return errors


def check_duplicates_cross_sheet(
    sheets: list[tuple[str, list, list[dict]]],
) -> list[tuple[str, int, str]]:
    """
    Check 7b (ERROR): cross-sheet duplicate key check.
    Intra-sheet duplicates are handled by check_duplicates. This function only
    reports cases where the same runtime key appears across different sheets.
    Returns: [(sheet_name, row_num, msg), ...]
    """
    errors: list[tuple[str, int, str]] = []
    exact_keys: dict[tuple, list] = {}
    literal_global: dict[str, list] = {}
    scoped_pattern_keys: dict[tuple, list] = {}
    pattern_global: dict[str, list] = {}
    substr_global: dict[str, list] = {}

    for sheet_name, _header, rows in sheets:
        for i, row in enumerate(rows, start=2):
            effective = _effective_method(row)
            untranslatable = row.get("untranslatable", "").strip().lower() in {"1", "true"}
            if effective == "ignore" or untranslatable:
                continue

            text = row.get("text", "")
            location = row.get("location", "").strip()
            key_5 = (
                location,
                row.get("parent", "").strip(),
                row.get("name", "").strip(),
                row.get("type", "").strip(),
                text,
            )

            if effective == "static":
                exact_keys.setdefault(key_5, []).append((sheet_name, i, row))
            elif effective == "literal":
                if location:
                    exact_keys.setdefault(key_5, []).append((sheet_name, i, row))
                else:
                    literal_global.setdefault(text, []).append((sheet_name, i, row))
            elif effective == "pattern":
                if location:
                    scoped_pattern_keys.setdefault(key_5, []).append((sheet_name, i, row))
                else:
                    pattern_global.setdefault(text, []).append((sheet_name, i, row))
            elif effective == "substr":
                substr_global.setdefault(text, []).append((sheet_name, i, row))

    def _emit(label: str, store: dict, is_tuple_key: bool):
        for key, occurrences in store.items():
            sheet_set = {sn for sn, _, _ in occurrences}
            if len(occurrences) <= 1 or len(sheet_set) <= 1:
                # intra-sheet duplicates are handled by check_duplicates
                continue
            locs = ", ".join(f"{sn}:{n}" for sn, n, _ in occurrences)
            if is_tuple_key:
                loc, parent, name, type_, text = key
                text_preview = _preview(text, 40)
                msg = (
                    f"{label} (cross-sheet, {len(occurrences)} occurrences): "
                    f"location={loc!r}, parent={parent!r}, name={name!r}, type={type_!r}, "
                    f"text={text_preview} [{locs}]"
                )
            else:
                text_preview = _preview(key, 40)
                msg = (
                    f"{label} (cross-sheet, {len(occurrences)} occurrences): "
                    f"text={text_preview} [{locs}]"
                )
            for sn, row_num, _ in occurrences:
                errors.append((sn, row_num, msg))

    _emit("exact match duplicate (static/scoped literal)", exact_keys, True)
    _emit("global literal duplicate", literal_global, False)
    _emit("scoped pattern duplicate", scoped_pattern_keys, True)
    _emit("global pattern duplicate", pattern_global, False)
    _emit("global substr duplicate", substr_global, False)

    return errors


# ==========================================
# output (Tee)
# ==========================================

class Tee:
    """Write print output to both screen and file."""

    def __init__(self, log_path: Path):
        self._log_path = log_path
        self._log_path.parent.mkdir(parents=True, exist_ok=True)
        self._fp = open(log_path, "w", encoding="utf-8")

    def print(self, *args, **kwargs):
        end = kwargs.get("end", "\n")
        sep = kwargs.get("sep", " ")
        text = sep.join(str(a) for a in args) + end
        print(*args, **kwargs)
        self._fp.write(text)

    def close(self):
        self._fp.close()


class ValidationResult:
    """Validation result. Passes when error_count is 0."""

    def __init__(self):
        self.error_tsv = 0
        self.error_flags = 0
        self.error_dup = 0
        self.error_method = 0
        self.warn_ws = 0
        self.warn_empty_method = 0
        self.warn_tsv_soft = 0
        self.log_path: Path | None = None

    @property
    def error_count(self) -> int:
        return (
            self.error_tsv + self.error_flags + self.error_dup + self.error_method
        )

    @property
    def warning_count(self) -> int:
        return self.warn_ws + self.warn_empty_method + self.warn_tsv_soft

    @property
    def ok(self) -> bool:
        return self.error_count == 0


# ==========================================
# run validation
# ==========================================

REQUIRED_COLUMNS = [
    "method", "filename", "filetype", "location",
    "parent", "name", "type", "unique_id",
    "text", "translation",
    "Transliteration", "Machine translated", "Confused", "untranslatable",
]


def validate_xlsx(xlsx_path: Path, tsv_dir: Path, soft: bool = False) -> ValidationResult:
    """
    Validate all translation sheets (excluding MetaData) in Translation.xlsx.

    soft=False (--hard, default): TSV match failure → ERROR (block build)
    soft=True  (--soft):          TSV match failure → WARNING (continue build, exclude that row)
    """
    result = ValidationResult()

    if not xlsx_path.exists():
        raise FileNotFoundError(f"xlsx file not found: {xlsx_path}")
    if not tsv_dir.exists():
        raise FileNotFoundError(
            f"TSV directory not found: {tsv_dir}\nRun parse_tscn_text.py first."
        )

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = xlsx_path.parent / ".log" / f"validate_translation_{timestamp}.log"
    result.log_path = log_path
    tee = Tee(log_path)

    try:
        tee.print(f"Target:    {xlsx_path}")
        tee.print(f"TSV source: {tsv_dir}")
        tee.print(f"Log file:  {log_path}")
        tee.print(f"Run time:  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        meta = load_metadata(xlsx_path)
        for line in format_metadata_lines(meta):
            tee.print(line)
        tee.print()

        tee.print("Building TSV index...")
        tsv_index = load_tsv_index(tsv_dir)
        total_tsv_entries = sum(len(v) for v in tsv_index.values())
        tee.print(f"  unique_id count: {len(tsv_index)}, total records: {total_tsv_entries}")

        tres_texts = load_tres_text_set(tsv_dir)
        gd_texts = load_gd_text_set(tsv_dir)
        tee.print(f"  tres text count: {len(tres_texts)}, gd text count: {len(gd_texts)}")

        tee.print("Loading Translation.xlsx...")
        sheets = load_all_translation_sheets(xlsx_path)
        if not sheets:
            tee.print("[ERROR] No translation sheets (no sheets other than MetaData)")
            raise ValueError("No translation sheets")
        total_row_count = sum(len(rows) for _, _, rows in sheets)
        tee.print(f"  {len(sheets)} sheets, {total_row_count} data rows")
        for sheet_name, _, rows in sheets:
            tee.print(f"    {sheet_name}: {len(rows)} rows")
        tee.print()

        for sheet_name, header, rows in sheets:
            missing = [c for c in REQUIRED_COLUMNS if c not in header]
            if missing:
                tee.print(f"[ERROR] [{sheet_name}] missing required columns: {missing}")
                tee.print(f"  Actual header: {header}")
                raise ValueError(f"[{sheet_name}] missing required columns: {missing}")

            # collect rows for per-sheet duplicate key check
            issues_by_row: dict = {}

            for i, row in enumerate(rows, start=2):
                local_issues = []

                for msg in check_tsv_match(row, tsv_index):
                    if soft:
                        local_issues.append(("WARN", "TSV match (soft)", msg))
                        result.warn_tsv_soft += 1
                    else:
                        local_issues.append(("ERROR", "TSV match", msg))
                        result.error_tsv += 1

                for msg in check_tres_text(row, tres_texts):
                    if soft:
                        local_issues.append(("WARN", "tres match (soft)", msg))
                        result.warn_tsv_soft += 1
                    else:
                        local_issues.append(("ERROR", "tres match", msg))
                        result.error_tsv += 1

                for msg in check_gd_text(row, gd_texts):
                    local_issues.append(("WARN", "gd match", msg))
                    result.warn_tsv_soft += 1

                for msg in check_whitespace(row):
                    local_issues.append(("WARN", "whitespace", msg))
                    result.warn_ws += 1

                for msg in check_flags(row):
                    local_issues.append(("ERROR", "flag", msg))
                    result.error_flags += 1

                for msg in check_method_fields(row):
                    local_issues.append(("ERROR", "method", msg))
                    result.error_method += 1

                for msg in check_empty_method(row):
                    local_issues.append(("WARN", "empty method", msg))
                    result.warn_empty_method += 1

                if local_issues:
                    issues_by_row.setdefault(i, []).extend(local_issues)

            # intra-sheet duplicate key check
            for row_num, msg in check_duplicates(rows):
                issues_by_row.setdefault(row_num, []).append(("ERROR", "duplicate key", msg))
                result.error_dup += 1

            if issues_by_row:
                tee.print(f"[{sheet_name}]")
                for i in sorted(issues_by_row.keys()):
                    row = rows[i - 2]
                    text_preview = _preview(row.get("text", ""), 60)
                    tee.print(f"  Row {i}: text={text_preview}")
                    for level, label, msg in issues_by_row[i]:
                        tee.print(f"    [{level}] {label}: {msg}")
                tee.print()

        # cross-sheet duplicate key check (same runtime key spans multiple sheets)
        cross_dup_by_sheet: dict = {}
        for sn, row_num, msg in check_duplicates_cross_sheet(sheets):
            cross_dup_by_sheet.setdefault(sn, {}).setdefault(row_num, []).append(msg)
            result.error_dup += 1
        if cross_dup_by_sheet:
            tee.print("[Cross-sheet duplicates]")
            sheet_rows_map = {sn: rows for sn, _, rows in sheets}
            for sn in sorted(cross_dup_by_sheet.keys()):
                tee.print(f"  [{sn}]")
                for i in sorted(cross_dup_by_sheet[sn].keys()):
                    row = sheet_rows_map[sn][i - 2]
                    text_preview = _preview(row.get("text", ""), 60)
                    tee.print(f"    Row {i}: text={text_preview}")
                    for msg in cross_dup_by_sheet[sn][i]:
                        tee.print(f"      [ERROR] duplicate key: {msg}")
            tee.print()

        tee.print("=" * 60)
        tee.print(f"Validation complete: {total_row_count} rows checked (mode={'soft' if soft else 'hard'})")
        tee.print(
            f"  ERROR {result.error_count}  "
            f"(TSV match {result.error_tsv}, flags {result.error_flags}, "
            f"duplicate keys {result.error_dup}, method {result.error_method})"
        )
        tee.print(
            f"  WARN  {result.warning_count}  "
            f"(whitespace {result.warn_ws}, empty method {result.warn_empty_method}"
            f"{f', TSV soft {result.warn_tsv_soft}' if result.warn_tsv_soft else ''})"
        )

        return result
    finally:
        tee.close()


def main() -> int:
    # parse --soft / --hard
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = [a for a in sys.argv[1:] if a.startswith("--")]
    soft = "--soft" in flags

    if not args:
        print("Usage: python validate_translation.py <locale> [--soft|--hard]")
        print("  --hard (default): TSV match failure -> ERROR (block build)")
        print("  --soft:           TSV match failure -> WARNING (continue build)")
        print("Example: python validate_translation.py Korean --soft")
        return 1

    locale = args[0]
    script_dir = Path(__file__).resolve().parent
    mod_root = script_dir.parent
    pkg_root = mod_root / "Trans To Vostok"
    xlsx_path = (pkg_root / locale / "Translation.xlsx").resolve()
    tsv_dir = (mod_root / ".tmp" / "parsed_text").resolve()

    if not xlsx_path.exists():
        print(f"[ERROR] xlsx file not found: {xlsx_path}")
        return 1

    try:
        result = validate_xlsx(xlsx_path, tsv_dir, soft=soft)
    except (FileNotFoundError, ValueError) as e:
        print(f"[ERROR] {e}")
        return 1

    return 0 if result.ok else 1


if __name__ == "__main__":
    sys.exit(main())
