"""Validate Texture canonical TSV (or xlsx) against parsed_textures/textures.tsv.

`parsed_textures/textures.tsv` (output of `tools/parse/parse_textures.py`) is
the PCK source of truth: every recovered PNG is cataloged with its size +
sha256. This validator joins each Texture row by (File Directory, File Name)
and reports inconsistencies.

Categories:

  ERRORS (exit code 1):
    ORPHAN                      row's (File Dir, File Name) absent from PCK
                                → texture removed from game; drop the row
    INVALID_METHOD              `method` value not in
                                {replace, blend, ignore, pending, ""}
    REPLACE_OR_BLEND_NO_TEXT    method=replace/blend with empty Text

  WARNINGS (exit code 2 if no errors):
    MISSING                     PCK has a texture not yet rowed in xlsx/TSV
                                → new texture; needs triage
    STALE_SIZE / STALE_SHA256   row's size/sha256 ≠ PCK's
                                → texture content changed; review overlay
                                  alignment (especially method=blend rows)
    LOCALE_NO_OVERLAY           (locale only) method=blend but
                                Translations/<locale>/textures/<path> missing
    LOCALE_NO_TRANSLATION       (locale only) method=replace/blend with
                                empty Translation cell

Schema checks (ORPHAN / INVALID_METHOD / REPLACE_OR_BLEND_NO_TEXT /
STALE_*) run for every locale including Template; the LOCALE_* checks
require an actual translation locale and are skipped for Template.

Usage:
    python tools/validate_texture.py Template
    python tools/validate_texture.py Korean
    python tools/validate_texture.py all
    python tools/validate_texture.py Korean --xlsx     # read .xlsx directly
    python tools/validate_texture.py all --report      # also write .tmp/ TSV reports
"""
import argparse
import csv
import sys
from pathlib import Path

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass


SCRIPT_DIR = Path(__file__).resolve().parent
REPO = SCRIPT_DIR.parent
TRANSLATIONS = REPO / "Translations"
PARSED_TEXTURES_TSV = REPO / ".tmp" / "parsed_textures" / "textures.tsv"
TEMPLATE_LOCALE = "Template"
CATEGORY = "Texture"

VALID_METHODS = {"replace", "blend", "ignore", "pending"}


def _norm_key(file_dir: str, file_name: str) -> str:
    """Same path normalization the rest of the project uses for Texture rows."""
    fd = (file_dir or "").strip().replace("/", "\\").strip("\\")
    fn = (file_name or "").strip()
    if fn:
        return (fd + "\\" + fn) if fd else fn
    return fd


def load_pck_catalog() -> dict[str, tuple[int, str]]:
    """Return {key -> (size, sha256)} from parsed_textures.tsv."""
    out: dict[str, tuple[int, str]] = {}
    with open(PARSED_TEXTURES_TSV, encoding="utf-8") as f:
        for row in csv.DictReader(f, delimiter="\t"):
            key = _norm_key(row["File Directory"], row["File Name"])
            out[key] = (int(row["size"]), row["sha256"])
    return out


def load_rows_from_tsv(locale: str) -> list[tuple[str, int, dict]]:
    """Return [(sheet_name, row_num, row_dict), ...] from canonical Texture TSVs."""
    out: list[tuple[str, int, dict]] = []
    tsv_dir = TRANSLATIONS / locale / "tsv" / CATEGORY
    if not tsv_dir.exists():
        return out
    for tsv in sorted(tsv_dir.glob("*.tsv")):
        if tsv.name.startswith("_"):
            continue
        with open(tsv, encoding="utf-8", newline="") as f:
            reader = csv.DictReader(f, delimiter="\t")
            for r_num, row in enumerate(reader, start=2):
                if not (row.get("File Directory") or row.get("File Name")):
                    continue
                out.append((tsv.stem, r_num, row))
    return out


def load_rows_from_xlsx(locale: str) -> list[tuple[str, int, dict]]:
    """Return [(sheet_name, row_num, row_dict), ...] from Texture.xlsx."""
    from openpyxl import load_workbook
    out: list[tuple[str, int, dict]] = []
    xlsx = TRANSLATIONS / locale / f"{CATEGORY}.xlsx"
    if not xlsx.exists():
        return out
    wb = load_workbook(xlsx, read_only=True, data_only=True)
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue
        header = [str(c or "").strip() for c in rows[0]]
        for r_num, r in enumerate(rows[1:], start=2):
            row_dict = {h: ("" if v is None else str(v))
                        for h, v in zip(header, r)}
            if not (row_dict.get("File Directory") or row_dict.get("File Name")):
                continue
            out.append((sheet, r_num, row_dict))
    return out


def validate_locale(locale: str, *, use_xlsx: bool, pck: dict[str, tuple[int, str]]
                    ) -> list[tuple[str, str, str, int, str, str]]:
    """Returns issues: [(severity, kind, sheet, row_num, key, detail), ...]."""
    is_template = (locale == TEMPLATE_LOCALE)
    rows = load_rows_from_xlsx(locale) if use_xlsx else load_rows_from_tsv(locale)
    issues: list[tuple[str, str, str, int, str, str]] = []
    seen_keys: set[str] = set()

    for sheet, r_num, row in rows:
        key = _norm_key(row.get("File Directory", ""), row.get("File Name", ""))
        if not key:
            continue
        seen_keys.add(key)
        method = (row.get("method") or "").strip().lower()
        text = (row.get("Text") or "").strip()

        # 1. ORPHAN
        if key not in pck:
            issues.append(("ERROR", "ORPHAN", sheet, r_num, key, ""))
            continue
        pck_size, pck_sha = pck[key]

        # 2. INVALID_METHOD (empty is allowed = untriaged)
        if method and method not in VALID_METHODS:
            issues.append(("ERROR", "INVALID_METHOD", sheet, r_num, key, method))

        # 3. REPLACE_OR_BLEND_NO_TEXT
        if method in ("replace", "blend") and not text:
            issues.append(("ERROR", "REPLACE_OR_BLEND_NO_TEXT", sheet, r_num, key, ""))

        # 4. STALE_SIZE
        tsv_size_raw = (row.get("size") or "").strip()
        if tsv_size_raw:
            try:
                tsv_size = int(tsv_size_raw)
            except ValueError:
                tsv_size = None
            if tsv_size is not None and tsv_size != pck_size:
                issues.append(("WARN", "STALE_SIZE", sheet, r_num, key,
                               f"{tsv_size} vs {pck_size}"))

        # 5. STALE_SHA256
        tsv_sha = (row.get("sha256") or "").strip()
        if tsv_sha and tsv_sha != pck_sha:
            issues.append(("WARN", "STALE_SHA256", sheet, r_num, key,
                           f"{tsv_sha[:8]} vs {pck_sha[:8]}"))

        # Locale-only checks
        if not is_template:
            translation = (row.get("Translation") or "").strip()
            if method in ("replace", "blend") and not translation:
                issues.append(("WARN", "LOCALE_NO_TRANSLATION", sheet, r_num, key, ""))
            if method == "blend":
                fd = (row.get("File Directory") or "").strip().replace("\\", "/").strip("/")
                fn = (row.get("File Name") or "").strip()
                if fd and fn:
                    overlay = TRANSLATIONS / locale / "textures" / fd / fn
                    if not overlay.exists():
                        issues.append(("WARN", "LOCALE_NO_OVERLAY", sheet, r_num, key,
                                       str(overlay.relative_to(REPO))))

    # 6. MISSING (PCK side)
    for key in sorted(pck):
        if key not in seen_keys:
            issues.append(("WARN", "MISSING", "", 0, key, ""))

    return issues


def report_console(locale: str, issues: list) -> tuple[int, int]:
    """Print a short summary; return (errors, warns) counts."""
    counts: dict[str, int] = {}
    for sev, kind, *_ in issues:
        k = f"{sev:5s} {kind}"
        counts[k] = counts.get(k, 0) + 1

    print(f"\n=== {locale} ===")
    if not counts:
        print("  OK (no issues)")
        return 0, 0

    errors = sum(c for k, c in counts.items() if k.startswith("ERROR"))
    warns = sum(c for k, c in counts.items() if k.startswith("WARN"))

    for k in sorted(counts):
        print(f"  {k:38s}  {counts[k]}")

    # First few samples per kind
    seen_kind: dict[str, int] = {}
    print("  Samples:")
    for sev, kind, sheet, r_num, key, detail in issues:
        if seen_kind.get(kind, 0) >= 2:
            continue
        seen_kind[kind] = seen_kind.get(kind, 0) + 1
        loc = f"{sheet}:r{r_num}" if sheet else ""
        det = f" ({detail})" if detail else ""
        print(f"    [{sev}] {kind:25s} {loc:18s} {key}{det}")

    return errors, warns


def write_report(locale: str, issues: list) -> Path:
    out = REPO / ".tmp" / f"texture_validation_{locale}.tsv"
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", encoding="utf-8", newline="") as f:
        w = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)
        w.writerow(["severity", "kind", "sheet", "row", "key", "detail"])
        for issue in issues:
            w.writerow(issue)
    return out


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("locale", help="Locale name or 'all'")
    parser.add_argument("--xlsx", action="store_true",
                        help="Read Texture.xlsx directly (default: canonical TSV)")
    parser.add_argument("--report", action="store_true",
                        help="Also write detailed TSV report under .tmp/")
    args = parser.parse_args(argv[1:])

    if not PARSED_TEXTURES_TSV.exists():
        print(f"[ERROR] PCK catalog not found: {PARSED_TEXTURES_TSV.relative_to(REPO)}")
        print(f"        Run: python tools/parse/parse_textures.py")
        return 1

    pck = load_pck_catalog()
    print(f"PCK catalog: {len(pck)} textures")

    if args.locale == "all":
        locales = sorted(p.name for p in TRANSLATIONS.iterdir() if p.is_dir())
    else:
        locales = [args.locale]

    total_errors = total_warns = 0
    for locale in locales:
        issues = validate_locale(locale, use_xlsx=args.xlsx, pck=pck)
        errors, warns = report_console(locale, issues)
        total_errors += errors
        total_warns += warns
        if args.report and issues:
            out = write_report(locale, issues)
            print(f"  Report: {out.relative_to(REPO)}")

    print(f"\n=== Grand Total ===")
    print(f"  Errors: {total_errors}")
    print(f"  Warns : {total_warns}")
    if total_errors:
        return 1
    if total_warns:
        return 2
    return 0


if __name__ == "__main__":
    sys.exit(main(sys.argv))
