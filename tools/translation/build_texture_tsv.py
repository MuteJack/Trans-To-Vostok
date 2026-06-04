"""Convert Texture.xlsx to canonical TSVs for one locale.

xlsx -> TSV side of the editing/VCS workflow, restricted to the
Texture category. Companion file build_translation_tsv.py handles
Translation.xlsx symmetrically.

Path mapping:
    Translations/<locale>/Texture.xlsx
        ->
    Translations/<locale>/tsv/Texture/<sheet>.tsv

Behavior:
    - All sheets exported (including MetaData)
    - Rows where every cell is None are skipped (truly empty visual separators)
    - All cell values are converted to str (avoids Excel type ambiguity)
    - Atomic write per file (.tmp + rename)
    - Stale TSVs (sheets that no longer exist in xlsx) are removed
    - Excel temp files ("~$...") detected and warned about
    - `_sheet_order.txt` written alongside the per-sheet TSVs so
      rebuild_texture_xlsx.py can preserve sheet order.

Input convention:
    <locale>   single locale
    Template   Template's Texture.xlsx
    all        every locale with enabled=true (English/Template excluded)

Usage:
    python tools/translation/build_texture_tsv.py Korean
    python tools/translation/build_texture_tsv.py Template
    python tools/translation/build_texture_tsv.py all
"""
import argparse
import csv
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl is required. pip install -r tools/requirements.txt", file=sys.stderr)
    sys.exit(1)

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass


SCRIPT_DIR = Path(__file__).resolve().parent  # tools/translation
TOOLS_DIR = SCRIPT_DIR.parent
REPO = TOOLS_DIR.parent
TRANSLATIONS_ROOT = REPO / "Translations"

sys.path.insert(0, str(TOOLS_DIR))
from utils.locale_config import load_locales  # noqa: E402

CATEGORY = "Texture"
XLSX_NAME = f"{CATEGORY}.xlsx"
CANONICAL_SUBDIR = "tsv"
TEMPLATE_LOCALE = "Template"
SOURCE_LOCALE = "English"
SKIP_FILE_PREFIXES = ("~$",)


def _say(msg: str = "") -> None:
    print(msg, flush=True)


def _normalize_cell(value) -> str:
    if value is None:
        return ""
    return str(value)


def _is_empty_row(row: tuple) -> bool:
    return all(v is None for v in row)


def export_xlsx(xlsx_path: Path, out_dir: Path) -> tuple[int, int]:
    """Export each sheet of an xlsx as <out_dir>/<sheet>.tsv.

    Returns (sheet_count, total_data_rows).
    """
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        expected_tsvs: set[str] = set()
        sheet_count = 0
        total_rows = 0

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            tsv_name = f"{sheet_name}.tsv"
            expected_tsvs.add(tsv_name)
            tsv_path = out_dir / tsv_name
            tmp = tsv_path.with_suffix(tsv_path.suffix + ".tmp")

            try:
                with open(tmp, "w", encoding="utf-8", newline="") as f:
                    writer = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)
                    sheet_rows = 0
                    for row_values in ws.iter_rows(values_only=True):
                        if row_values is None:
                            continue
                        if _is_empty_row(row_values):
                            continue
                        cells = [_normalize_cell(v) for v in row_values]
                        writer.writerow(cells)
                        sheet_rows += 1
                tmp.replace(tsv_path)
                sheet_count += 1
                total_rows += sheet_rows
                _say(f"    -> {tsv_name}  ({sheet_rows} rows)")
            except Exception:
                if tmp.exists():
                    try:
                        tmp.unlink()
                    except OSError:
                        pass
                raise

        for stale in out_dir.glob("*.tsv"):
            if stale.name not in expected_tsvs:
                try:
                    stale.unlink()
                    _say(f"    x removed stale: {stale.name}")
                except OSError as e:
                    _say(f"    [WARN] Failed to remove stale: {stale.name} ({e})")

        order_path = out_dir / "_sheet_order.txt"
        order_tmp = order_path.with_suffix(order_path.suffix + ".tmp")
        try:
            with open(order_tmp, "w", encoding="utf-8", newline="\n") as f:
                for name in wb.sheetnames:
                    f.write(name + "\n")
            order_tmp.replace(order_path)
            _say(f"    -> _sheet_order.txt ({len(wb.sheetnames)} sheets)")
        except Exception:
            if order_tmp.exists():
                try:
                    order_tmp.unlink()
                except OSError:
                    pass
            raise

        stale_widths = out_dir / "_column_widths.json"
        if stale_widths.exists():
            try:
                stale_widths.unlink()
                _say("    x removed stale: _column_widths.json")
            except OSError as e:
                _say(f"    [WARN] Failed to remove stale: _column_widths.json ({e})")

        return sheet_count, total_rows
    finally:
        wb.close()


def build_for_locale(locale: str) -> int:
    locale_dir = TRANSLATIONS_ROOT / locale
    xlsx_path = locale_dir / XLSX_NAME
    if not xlsx_path.exists():
        _say(f"  [SKIP] {xlsx_path.relative_to(REPO)} not found")
        return 0
    if xlsx_path.name.startswith(SKIP_FILE_PREFIXES):
        _say(f"  [SKIP] Excel lock file: {xlsx_path.name}")
        return 0

    out_dir = locale_dir / CANONICAL_SUBDIR / CATEGORY
    _say(f"  {xlsx_path.relative_to(REPO)} -> {out_dir.relative_to(REPO)}/")
    try:
        sc, rc = export_xlsx(xlsx_path, out_dir)
    except PermissionError:
        _say(f"  [ERROR] Cannot read xlsx (file open in Excel?): {xlsx_path}")
        return 1
    _say(f"  -> {sc} sheets, {rc} rows")
    return 0


def _discover_enabled_locales() -> list[str]:
    out: list[str] = []
    for loc in load_locales():
        name = loc.get("dir")
        if loc.get("enabled") and name and name not in (SOURCE_LOCALE, TEMPLATE_LOCALE):
            out.append(name)
    return out


def main(argv: list[str]) -> int:
    parser = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("locale", help="Locale name, 'Template', or 'all'")
    args = parser.parse_args(argv[1:])

    if args.locale == "all":
        locales = _discover_enabled_locales()
        if not locales:
            _say("[ERROR] no enabled locales found in src/locale.json")
            return 1
    else:
        if not (TRANSLATIONS_ROOT / args.locale).exists():
            _say(f"[ERROR] locale directory not found: Translations/{args.locale}")
            return 1
        locales = [args.locale]

    _say(f"=== build_texture_tsv ({len(locales)} target(s)) ===")
    _say(f"  Targets: {', '.join(locales)}")
    _say()

    overall = 0
    for loc in locales:
        _say(f"--- {loc} ---")
        rc = build_for_locale(loc)
        if rc != 0:
            overall = rc
        _say()

    return overall


if __name__ == "__main__":
    sys.exit(main(sys.argv))
