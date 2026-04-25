"""
Images.xlsx → Attribution.md 자동 생성.

각 시트의 다음 컬럼을 읽어 Markdown 으로 출력:
    - File Name
    - Reworked by
    - Attribution

사용법:
    python tools/build_attributions.py                       # 기본 (Korean → Attribution.md)
    python tools/build_attributions.py --locale Korean
    python tools/build_attributions.py --output custom.md
"""
import argparse
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl이 필요합니다. pip install openpyxl", file=sys.stderr)
    sys.exit(1)

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass


URL_RE = re.compile(r'https?://[^\s)\]]+')

REQUIRED_COLUMNS = ["File Name", "Reworked by", "Attribution"]


def collect_rows(xlsx_path: Path) -> list[dict]:
    """모든 시트에서 (sheet, file_name, reworked_by, attribution) 행 수집."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    rows: list[dict] = []
    for ws in wb.worksheets:
        header = [c.value for c in ws[1]]
        try:
            idx = {col: header.index(col) for col in REQUIRED_COLUMNS}
        except ValueError as e:
            print(f"[WARN] '{ws.title}' 시트에 필수 컬럼 누락 ({e}), 스킵",
                  file=sys.stderr)
            continue

        for row_no, row in enumerate(ws.iter_rows(values_only=True), 1):
            if row_no == 1:
                continue
            def cell(name: str) -> str:
                i = idx[name]
                return str(row[i]).strip() if i < len(row) and row[i] is not None else ""

            file_name = cell("File Name")
            if not file_name:
                continue  # 빈 행 스킵

            rows.append({
                "sheet": ws.title,
                "file_name": file_name,
                "reworked_by": cell("Reworked by"),
                "attribution": cell("Attribution"),
            })
    return rows


def linkify(text: str) -> str:
    """텍스트 내부의 URL 을 markdown auto-link 로 변환."""
    return URL_RE.sub(lambda m: f"<{m.group(0)}>", text)


def render_markdown(rows: list[dict], locale: str) -> str:
    with_attr = [r for r in rows if r["attribution"]]
    without_attr = [r for r in rows if not r["attribution"]]

    lines: list[str] = []
    lines.append(f"# Attribution — Trans To Vostok ({locale})")
    lines.append("")
    lines.append(
        "This document lists attributions for translated image assets in this mod."
    )
    lines.append("")
    lines.append(
        f"_Auto-generated from `{locale}/Images.xlsx` by "
        "`tools/build_attributions.py`. Do not edit manually — "
        "update the xlsx instead._"
    )
    lines.append("")

    if with_attr:
        lines.append("## Files with third-party attribution")
        lines.append("")
        for r in with_attr:
            lines.append(f"### `{r['file_name']}` _({r['sheet']})_")
            lines.append("")
            lines.append(f"- **Reworked by**: {r['reworked_by'] or '(unknown)'}")
            lines.append("- **Sources**:")
            for src_line in r["attribution"].splitlines():
                src_line = src_line.strip()
                if not src_line:
                    continue
                lines.append(f"  - {linkify(src_line)}")
            lines.append("")

    if without_attr:
        lines.append("## Files without third-party attribution")
        lines.append("")
        lines.append(
            "Created by the listed translator(s) without using third-party sources."
        )
        lines.append("")
        # 시트별 그룹
        by_sheet: dict[str, list[dict]] = {}
        for r in without_attr:
            by_sheet.setdefault(r["sheet"], []).append(r)
        for sheet_name in by_sheet:
            lines.append(f"### {sheet_name}")
            lines.append("")
            for r in by_sheet[sheet_name]:
                rework = r["reworked_by"] or "(unknown)"
                lines.append(f"- `{r['file_name']}` — {rework}")
            lines.append("")

    if not rows:
        lines.append("_No image assets registered._")
        lines.append("")

    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate Attribution.md from Images.xlsx"
    )
    parser.add_argument(
        "--locale", default="Korean",
        help="Locale folder name (default: Korean)",
    )
    parser.add_argument(
        "--output", default=None,
        help="Output file path (default: <pkg_root>/<locale>/Attribution.md)",
    )
    args = parser.parse_args()

    script_dir = Path(__file__).resolve().parent
    mod_root = script_dir.parent
    pkg_root = mod_root / "Trans To Vostok"
    xlsx_path = pkg_root / args.locale / "Images.xlsx"

    if not xlsx_path.exists():
        # 빌드 파이프라인에서 호출될 때 Images.xlsx 없는 로케일은 정상 (텍스처 번역 안 함)
        print(f"[SKIP] Images.xlsx 가 없습니다: {xlsx_path} — 출력 안 함")
        return 0

    print(f"Input:  {xlsx_path}")
    rows = collect_rows(xlsx_path)
    sheets = sorted({r['sheet'] for r in rows})
    print(f"Loaded: {len(rows)} entries from sheets: {', '.join(sheets) or '-'}")

    md = render_markdown(rows, args.locale)

    out_path = (
        Path(args.output).resolve()
        if args.output
        else (pkg_root / args.locale / "Attribution.md")
    )
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(md, encoding="utf-8")
    print(f"Output: {out_path}")

    with_attr = sum(1 for r in rows if r["attribution"])
    without_attr = len(rows) - with_attr
    print()
    print(f"  with attribution:    {with_attr}")
    print(f"  without attribution: {without_attr}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
