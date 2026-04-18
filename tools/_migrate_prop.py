"""
[임시 도구] xlsx 를 읽어 tres 행의 name 컬럼 값을 새 prop 컬럼으로 이동.

워크플로:
    1. 이 도구 실행 → .tmp/Translation_migrated.tsv 생성
    2. 결과 TSV 를 엑셀에 수동으로 붙여넣기

변환 규칙:
    filetype=tres 행:   prop = 기존 name 값 | name = 빈값
    filetype=tscn 행:   prop = "text" (기본) | name = 기존 그대로
    filetype=gd 행:     기존 그대로 (prop 추가만)
    기타:               기존 그대로 (prop 추가만)

출력 컬럼 순서 (기존 name/type 사이에 prop 삽입):
    WHERE, SUB, KIND, Transliteration, Machine Translated, Confused, ignore,
    method, filename, filetype, location, parent, name, type, prop, unique_id,
    text, translation, DESCRIPTION

사용법:
    python _migrate_prop.py                    # 기본: Korean/Translation.xlsx
    python _migrate_prop.py <xlsx_path>        # 지정
    python _migrate_prop.py <xlsx_path> <tsv_path>
"""
import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl이 필요합니다. pip install openpyxl", file=sys.stderr)
    sys.exit(1)

if sys.stdout.encoding and sys.stdout.encoding.lower() not in ("utf-8", "utf8"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except (AttributeError, Exception):
        pass


_X000D_RE = re.compile(r"_x000[dD]_")

# property 컬럼을 삽입할 위치 (type 뒤, unique_id 앞)
# 실제 컬럼 목록은 xlsx 에서 동적으로 읽음 (하드코딩 안 함)
PROP_INSERT_AFTER = "type"
PROP_COLUMN_NAME = "property"
SHEET_COLUMN_NAME = "_sheet"


def _normalize(value) -> str:
    if value is None:
        return ""
    s = str(value)
    return _X000D_RE.sub("", s)


def _normalize_header(value) -> str:
    s = _normalize(value)
    s = s.replace("\r", " ").replace("\n", " ")
    return " ".join(s.split())


def migrate_row(row: dict) -> dict:
    """
    한 행을 마이그레이션. 기존 row 를 수정하지 않고 새 dict 반환.

    tres 행: name 값을 prop 으로 이동, name 은 빈값
    tscn 행: prop 이 없으면 "text" 기본값 (name 은 그대로 Godot 노드명)
    기타:     prop 그대로 (없으면 빈값)
    """
    new_row = dict(row)  # 얕은 복사
    filetype = new_row.get("filetype", "").strip()
    existing_prop = new_row.get("property", "").strip()

    if filetype == "tres":
        # name 값이 있고 prop 이 비어있으면 이동
        if existing_prop:
            # 이미 prop 채워져 있음 — 건드리지 않음
            pass
        else:
            name_val = new_row.get("name", "").strip()
            new_row["property"] = name_val
            new_row["name"] = ""
    elif filetype == "tscn":
        if not existing_prop:
            new_row["property"] = "text"
    # gd / 기타 filetype 은 기존 prop 유지 (빈값 허용)

    return new_row


def process_sheet(ws, sheet_name: str) -> tuple[list[str], list[dict]]:
    """시트 하나를 읽어 (기존 헤더, 행 리스트) 반환."""
    rows_iter = ws.iter_rows(values_only=True)
    header_raw = next(rows_iter, None)
    if header_raw is None:
        return [], []
    header = [_normalize_header(c) for c in header_raw]

    rows = []
    for row_values in rows_iter:
        if row_values is None or all(v is None for v in row_values):
            continue
        row_dict: dict = {}
        for i, key in enumerate(header):
            val = row_values[i] if i < len(row_values) else None
            # 중복 헤더는 첫 번째만 사용
            if key not in row_dict:
                row_dict[key] = _normalize(val)
        rows.append(row_dict)
    return header, rows


def main() -> int:
    script_dir = Path(__file__).resolve().parent
    mod_root = script_dir.parent

    default_xlsx = mod_root / "Korean" / "Translation.xlsx"
    default_tsv = mod_root / ".tmp" / "Translation_migrated.tsv"

    xlsx_path = Path(sys.argv[1]).resolve() if len(sys.argv) > 1 else default_xlsx.resolve()
    tsv_path = Path(sys.argv[2]).resolve() if len(sys.argv) > 2 else default_tsv.resolve()

    if not xlsx_path.exists():
        print(f"[ERROR] xlsx 파일이 없습니다: {xlsx_path}", file=sys.stderr)
        return 1

    print(f"입력: {xlsx_path}")
    print(f"출력: {tsv_path}")
    print()

    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    try:
        all_rows: list[dict] = []
        sheet_stats: list[tuple[str, int, int]] = []  # (sheet, total, tres_migrated)
        reference_header: list[str] = []

        for sheet_name in wb.sheetnames:
            if sheet_name == "MetaData":
                continue
            ws = wb[sheet_name]
            header, rows = process_sheet(ws, sheet_name)

            # 첫 시트 헤더를 기준으로 삼음 (모든 시트가 같은 구조라고 가정)
            if not reference_header and header:
                reference_header = header

            tres_migrated = 0
            migrated_rows = []
            for row in rows:
                new_row = migrate_row(row)
                new_row[SHEET_COLUMN_NAME] = sheet_name  # 시트 이름 주입
                if (row.get("filetype", "").strip() == "tres"
                        and row.get("name", "").strip()
                        and not row.get("property", "").strip()):
                    tres_migrated += 1
                migrated_rows.append(new_row)

            all_rows.extend(migrated_rows)
            sheet_stats.append((sheet_name, len(rows), tres_migrated))

        # 최종 컬럼 순서: _sheet + reference_header (prop 삽입) + 누락분
        if not reference_header:
            print("[ERROR] 헤더를 읽을 수 없습니다.", file=sys.stderr)
            return 1

        # reference_header 에 prop 없으면 PROP_INSERT_AFTER 뒤에 삽입
        final_columns = [SHEET_COLUMN_NAME]
        prop_inserted = False
        for col in reference_header:
            final_columns.append(col)
            if col == PROP_COLUMN_NAME:
                prop_inserted = True
            elif col == PROP_INSERT_AFTER and PROP_COLUMN_NAME not in reference_header:
                final_columns.append(PROP_COLUMN_NAME)
                prop_inserted = True

        if not prop_inserted:
            # type 컬럼이 없었거나 prop 이 이미 어딘가에 있음 — 그래도 보장 차원에서 추가
            if PROP_COLUMN_NAME not in final_columns:
                final_columns.append(PROP_COLUMN_NAME)

        # TSV 출력
        tsv_path.parent.mkdir(parents=True, exist_ok=True)
        tmp_path = tsv_path.with_suffix(tsv_path.suffix + ".tmp")
        with open(tmp_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f, delimiter="\t", quoting=csv.QUOTE_MINIMAL)
            writer.writerow(final_columns)
            for row in all_rows:
                writer.writerow([row.get(c, "") for c in final_columns])
        tmp_path.replace(tsv_path)

        print(f"컬럼 ({len(final_columns)}): {final_columns}")
        print()

        # 요약
        print("시트별 처리 결과:")
        total_migrated = 0
        for sheet, total, migrated in sheet_stats:
            print(f"  {sheet}: {total}행, tres→prop 이동 {migrated}개")
            total_migrated += migrated
        print()
        print(f"전체: {len(all_rows)}행, tres→prop 이동 총 {total_migrated}개")
        print(f"출력: {tsv_path}")
        print()
        print("다음 단계: 이 TSV 의 데이터 영역을 복사해 엑셀의 새 스키마에 붙여넣기")
        print("(엑셀 컬럼 순서도 동일하게 맞춰야 함)")
        return 0
    finally:
        wb.close()


if __name__ == "__main__":
    sys.exit(main())
